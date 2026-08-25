#fichier qui crée une interface permettant de modifier des composants (composants électriques, clefs) des tableaux

import tkinter as tk
from tkinter import ttk, messagebox, PhotoImage
import pandas as pd
from openpyxl import load_workbook
from LHT_Classe_dessin import Dessin
from LHT_Classe_etapes_choisies import EtapesChoisies
from LHT_Classe_initialisation_cellules import InitialisationCellules
from Classe_executionLHT import ExecutionLHT
import ast

#la classe ModifsExcel crée une interface pour modifier des composants électrique plus facilement que sur l'Excel
class ModifsExcel:
    #caractéristiques de la classe
    def __init__(self, root):
        self.key_entries_list = []  #cette liste stocke les clefs ajoutées 
        self.window = tk.Toplevel(root)
        self.window.title("Ajouter des elements")
        icon = PhotoImage(file='Logo_sans_fond.png')
        self.window.iconphoto(False, icon)
        self.window.iconbitmap('Logo_avec_fond.ico')
        self.window.geometry("1400x600")

        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(expand=True, fill='both')
        #reprend la dernière valeur de palier choisie
        try:
            with open("dataLHT/dernier_palier_choisi_LHT.txt", "r") as file:
                self.palier = file.read().strip()
        except FileNotFoundError:
            self.palier = "900"  #si le fichier n'est pas trouvé, on met par défaut le palier 1400
        print(self.palier)
        self.cell_size = 30

        #crée chaque onglet avec les données nécessaires
        self.create_palier_tab()
        self.create_key_tab()
        self.create_coffret_tab()
        self.create_panneau_tab()
        #self.create_fusible_tab()
        self.create_transformateur_tab()
        self.create_serrure_tab()
        self.create_source_tab()
        self.create_clefs_libres_tab()
        self.create_partie_mobile_tab()
        self.create_eclisse_tab()
        self.create_boite_eclisse_tab()
        self.create_smalt_tab()
        self.create_cellule_tab()
        self.create_update_palier_tab()
        self.create_retour_palier_tab()

    #crée l'onglet permettant le choisir le palier sur lequel on travaille
    def create_palier_tab(self):
        self.palier_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.palier_tab, text="Sélectionner Palier")
        
        tk.Label(self.palier_tab, text="Choisissez un palier (900 ou 1300 2 tranches ou 1300 4 tranches ou 1400):").pack(padx=10, pady=10)
        
        #lit la dernière valeur de palier depuis le fichier
        try:
            with open("dataLHT/dernier_palier_choisi_LHT.txt", "r") as file:
                dernier_palier_choisi = file.read().strip()
        except FileNotFoundError:
            dernier_palier_choisi = "900"  #valeur par défaut si le fichier n'existe pas
        self.palier_var = tk.StringVar(value=dernier_palier_choisi)
        palier_menu = ttk.OptionMenu(self.palier_tab, self.palier_var, dernier_palier_choisi, "900", "1300_2tranches", "1300_4tranches", "1400")
        palier_menu.pack(padx=10, pady=10)
        
        submit_button = ttk.Button(self.palier_tab, text="Soumettre", command=self.set_palier)
        submit_button.pack(padx=10, pady=10)

        tableau_button = ttk.Button(self.palier_tab, text="Visualisation tableau", command=self.tableauLHT_tab)
        tableau_button.pack(padx=10, pady=10)

    #permet de valider la sélection du palier et de l'enregistrer pour repartir de ce palier les fois suivantes
    def set_palier(self):
        self.palier = self.palier_var.get()
        with open("dataLHT/dernier_palier_choisi_LHT.txt", "w") as file:
            file.write(self.palier)
        
        # Recharge le tableau
        self.load_and_draw_all_cells()
        self.canvas.move("all", 0, 15)
        
        # Supprimer et recréer l'onglet des clefs si il existe
        if hasattr(self, 'key_tab'):
            key_tab_index = self.notebook.index(self.key_tab)
            self.notebook.forget(key_tab_index)
            del self.key_tab
        self.create_key_tab()

        # Supprimer et recréer l'onglet des coffrets si il existe
        if hasattr(self, 'coffret_tab'):
            coffret_tab_index = self.notebook.index(self.coffret_tab)
            self.notebook.forget(coffret_tab_index)
            del self.coffret_tab
        self.create_coffret_tab()

        '''
        # Supprimer et recréer l'onglet des fusibles si il existe
        if hasattr(self, 'fusible_tab'):
            fusible_tab_index = self.notebook.index(self.fusible_tab)
            self.notebook.forget(fusible_tab_index)
            del self.fusible_tab
        self.create_fusible_tab()
        '''

        # Supprimer et recréer l'onglet des panneaux si il existe
        if hasattr(self, 'panneau_tab'):
            panneau_tab_index = self.notebook.index(self.panneau_tab)
            self.notebook.forget(panneau_tab_index)
            del self.panneau_tab
        self.create_panneau_tab()

        # Supprimer et recréer l'onglet des transformateurs si il existe
        if hasattr(self, 'transformateur_tab'):
            transformateur_tab_index = self.notebook.index(self.transformateur_tab)
            self.notebook.forget(transformateur_tab_index)
            del self.transformateur_tab
        self.create_transformateur_tab()

        # Supprimer et recréer l'onglet des serrures si il existe
        if hasattr(self, 'serrure_tab'):
            serrure_tab_index = self.notebook.index(self.serrure_tab)
            self.notebook.forget(serrure_tab_index)
            del self.serrure_tab
        self.create_serrure_tab()

        # Supprimer et recréer l'onglet des Sources si il existe
        if hasattr(self, 'source_tab'):
            source_tab_index = self.notebook.index(self.source_tab)
            self.notebook.forget(source_tab_index)
            del self.source_tab
        self.create_source_tab()

        # Supprimer et recréer l'onglet des clefs libres si il existe
        if hasattr(self, 'clefs_libres_tab'):
            clefs_libres_tab_index = self.notebook.index(self.clefs_libres_tab)
            self.notebook.forget(clefs_libres_tab_index)
            del self.clefs_libres_tab
        self.create_clefs_libres_tab()

        # Supprimer et recréer l'onglet des parties mobiles si il existe
        if hasattr(self, 'partie_mobile_tab'):
            partie_mobile_tab_index = self.notebook.index(self.partie_mobile_tab)
            self.notebook.forget(partie_mobile_tab_index)
            del self.partie_mobile_tab
        self.create_partie_mobile_tab()

        # Supprimer et recréer l'onglet des Eclisses si il existe
        if hasattr(self, 'eclisse_tab'):
            eclisse_tab_index = self.notebook.index(self.eclisse_tab)
            self.notebook.forget(eclisse_tab_index)
            del self.eclisse_tab
        self.create_eclisse_tab()

        # Supprimer et recréer l'onglet des Boites à Eclisse si il existe
        if hasattr(self, 'boite_eclisse_tab'):
            boite_eclisse_tab_index = self.notebook.index(self.boite_eclisse_tab)
            self.notebook.forget(boite_eclisse_tab_index)
            del self.boite_eclisse_tab
        self.create_boite_eclisse_tab()

        # Supprimer et recréer l'onglet des smalts si il existe
        if hasattr(self, 'smalt_tab'):
            smalt_tab_index = self.notebook.index(self.smalt_tab)
            self.notebook.forget(smalt_tab_index)
            del self.smalt_tab
        self.create_smalt_tab()

        # Supprimer et recréer l'onglet des cellules si il existe
        if hasattr(self, 'cellule_tab'):
            cellule_tab_index = self.notebook.index(self.cellule_tab)
            self.notebook.forget(cellule_tab_index)
            del self.cellule_tab
        self.create_cellule_tab()

        # Supprimer et recréer l'onglet pour mettre à jour le ficher Excel si il existe
        if hasattr(self, 'update_palier_tab'):
            update_palier_tab_index = self.notebook.index(self.update_palier_tab)
            self.notebook.forget(update_palier_tab_index)
            del self.update_palier_tab
        self.create_update_palier_tab()

        # Supprimer et recréer l'onglet pour revenir à l'état avant les modifications si il existe
        if hasattr(self, 'retour_palier_tab'):
            retour_palier_tab_index = self.notebook.index(self.retour_palier_tab)
            self.notebook.forget(retour_palier_tab_index)
            del self.retour_palier_tab
        self.create_retour_palier_tab()

    #crée l'onglet permettant de modifier des clefs
    def create_key_tab(self):
        self.key_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.key_tab, text="Clés")
        
        self.key_labels = ["nom", "etat", "penne", "appartenance", "type", "cellule"]
        self.key_entries = {}

        # Récupération des valeurs uniques pour 'nom', 'appartenance' et 'cellule' dans les données
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        book = load_workbook(excel_path)
        sheet_name = "clef"
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            noms = df['nom'].unique().tolist()
            appartenance_values = df['appartenance'].unique().tolist()
            cellules = df['cellule'].unique().tolist()
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        # Création des champs
        for i, label in enumerate(self.key_labels):
            if label == "penne":
                ttk.Label(self.key_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
                self.penne_var = tk.StringVar(value="rentre")
                penne_menu = ttk.OptionMenu(self.key_tab, self.penne_var, "rentre", "rentre", "sorti", "aucune")
                penne_menu.grid(row=i, column=1, padx=10, pady=5)
            
            elif label == "etat":
                ttk.Label(self.key_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
                self.key_etat_var = tk.StringVar(value="prisonniere")
                etat_menu = ttk.OptionMenu(self.key_tab, self.key_etat_var, "prisonniere", "prisonniere", "presente", "absente")
                etat_menu.grid(row=i, column=1, padx=10, pady=5)
            
            elif label == "type":
                ttk.Label(self.key_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
                self.key_type_var = tk.StringVar(value="serrure_mere")
                etat_menu = ttk.OptionMenu(self.key_tab, self.key_type_var, "serrure_mere", "serrure_mere", "partie_mobile", "smalt", "coffret", "transformateur", "panneau", "eclisse", "boite_eclisse")
                etat_menu.grid(row=i, column=1, padx=10, pady=5)
                ttk.Label(self.key_tab, text="(écrire le type d'élément électrique auquel la clé appartient)").grid(row=i, column=2, padx=10, pady=5)

            else:
                ttk.Label(self.key_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
                
                if label == "nom":
                    # Créer un combobox pour 'nom' avec les valeurs uniques récupérées
                    combobox = ttk.Combobox(self.key_tab, values=noms)
                    combobox.grid(row=i, column=1, padx=10, pady=5)
                    self.key_entries[label] = combobox
                    
                    ttk.Label(self.key_tab, text="(Appuyez sur la touche Entrée si vous voulez rechercher une clef portant ce nom)").grid(row=i, column=2, padx=10, pady=5)
                    combobox.bind("<Return>", self.check_key_existence)

                elif label == "appartenance":
                    # Créer un combobox pour 'appartenance' avec les valeurs uniques récupérées
                    combobox = ttk.Combobox(self.key_tab, values=appartenance_values)
                    combobox.grid(row=i, column=1, padx=10, pady=5)
                    self.key_entries[label] = combobox
                    ttk.Label(self.key_tab, text="(écrire le nom de l'élément électrique auquel la clé appartient)").grid(row=i, column=2, padx=10, pady=5)

                elif label == "cellule":
                    # Créer un combobox pour 'cellule' avec les valeurs uniques récupérées
                    combobox = ttk.Combobox(self.key_tab, values=cellules)
                    combobox.grid(row=i, column=1, padx=10, pady=5)
                    self.key_entries[label] = combobox
                    ttk.Label(self.key_tab, text="(écrire le nom de la cellule auquel la clé appartient)").grid(row=i, column=2, padx=10, pady=5)
                    
        # Ajout des boutons de sauvegarde et suppression
        self.save_key_button = ttk.Button(self.key_tab, text="Sauvegarder", command=self.save_keys)
        self.save_key_button.grid(row=len(self.key_labels), columnspan=2, padx=10, pady=10)
        
        self.delete_key_button = ttk.Button(self.key_tab, text="Supprimer", command=self.delete_key)
        self.delete_key_button.grid(row=len(self.key_labels) + 1, columnspan=2, padx=10, pady=10)
        
        # Variable pour stocker l'index de la ligne sélectionnée
        self.selected_row_index = None

    #vérifie si des clefs avec ce nom existent déjà    
    def check_key_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.key_entries["nom"].get()  # Utilise le combobox pour 'nom'
        rows = df[df['nom'] == nom_to_load]
        
        if not rows.empty:
            print("les lignes contenant la clef ne sont pas vides")
            choix = self.choisir_ligne(rows)
            if choix is None:
                self.selected_row_index = None
                return
            row = rows.iloc[choix]
            self.selected_row_index = row.name
            self.key_etat_var.set(row["etat"])
            self.penne_var.set(row["penne"])
            self.key_entries["appartenance"].set(row["appartenance"])  # Utilise le combobox pour 'appartenance'
            self.key_type_var.set(row["type"])
            self.key_entries["cellule"].set(row["cellule"])  # Utilise le combobox pour 'cellule'
        else:
            self.selected_row_index = None

    #nouvelle fenêtre qui affiche les différentes clefs possédant le nom écrit, pour pouvoir choisir une des clefs        
    def choisir_ligne(self, rows):
        dialog = tk.Toplevel(self.window)
        dialog.title("Choix de la ligne")
        
        tk.Label(dialog, text=f"Une ou plusieurs lignes correspondent au nom '{self.key_entries['nom'].get()}'. \nVous pouvez soit choisir une clef et modifier ses caracteristiques, soit simplement fermer la fenetre et créer une nouvelle clef.").pack(pady=10)

        listbox = tk.Listbox(dialog, width=130)
        lignes = []
        for index, row in rows.iterrows():
            listbox.insert(tk.END, f"Ligne {index + 1}: {row.to_dict()}")
            lignes.append(index)
        listbox.pack(pady=10)
        print("lignes", lignes)
        dialog.selected_index = None

        #fonction qui charge les caractéristiques de la clef choisie
        def on_select():
            selection = listbox.curselection()
            ligne = lignes[selection[0]]
            print("ligne", ligne)
            if ligne:
                dialog.selected_index = selection[0]
            else:
                dialog.selected_index = None
            dialog.destroy()
        
        select_button = ttk.Button(dialog, text="Sélectionner", command=on_select)
        select_button.pack(pady=10)
        
        dialog.transient(self.window)
        dialog.grab_set()
        self.window.wait_window(dialog)
        print()
        print("selected_index", dialog.selected_index)
        return getattr(dialog, 'selected_index', None)
    
    #fonction qui sauvergarde les modifications qui la clef avec les caractéristiques que l'on a écrit
    def save_keys(self):
        new_key = {label: self.key_entries[label].get() for label in self.key_labels if label not in ["penne", "etat", "type"]}
        new_key["penne"] = self.penne_var.get()
        new_key["etat"] = self.key_etat_var.get()
        new_key["type"] = self.key_type_var.get()
        
        #Ajoute la nouvelle clé à la liste des clés
        self.key_entries_list.append(new_key)
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")

        #ajoute la clef créée ou sauvegarde les modifications d'une clef existante, en mettant à jour les changements que cela impose
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        old_appartenance = None
        old_type = None
        if self.selected_row_index is not None:
            old_appartenance = df.at[self.selected_row_index, 'appartenance']
            old_type = df.at[self.selected_row_index, 'type']
            #Modifie la ligne existante
            df.loc[self.selected_row_index] = new_key
            print(new_key)
            print(f"Clé modifiée dans {self.palier}_modifs_LHT.xlsx, feuille 'clef'")
            self.selected_row_index = None
        else:
            #Ajoute une nouvelle ligne
            df = df.append(new_key, ignore_index=True)
            print(f"Nouvelles clés sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'clef'")

        
        book.remove(sheet)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        #met à jour la représentation de la serrurerie à laquelle la clef appartient dans le cas où elle appartient à un smalt ou une partie mobile
        if new_key["type"] in ["partie_mobile", "smalt", "boite_eclisse", "eclisse"]:
            if old_appartenance and old_appartenance != new_key["appartenance"]:
                self.update_representation(book, old_type, old_appartenance)
            self.update_representation(book, new_key["type"], new_key["appartenance"])
        
    #supprime une clef qui existe
    def delete_key(self):
        nom_to_delete = self.key_entries["nom"].get()
        etat_to_delete = self.key_etat_var.get()
        penne_to_delete = self.penne_var.get()
        appartenance_to_delete = self.key_entries["appartenance"].get()
        type_to_delete = self.key_type_var.get()
        cellule_to_delete = self.key_entries["cellule"].get()
        
        if not nom_to_delete or not appartenance_to_delete or not cellule_to_delete:
            messagebox.showwarning("Avertissement", "Veuillez remplir tout les champs")
            return
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        #Trouve les lignes correspondant aux valeurs des entrées
        rows_to_delete = df[
            (df['nom'] == nom_to_delete) &
            (df['etat'] == etat_to_delete) &
            (df['penne'] == penne_to_delete) &
            (df['appartenance'] == appartenance_to_delete) &
            (df['type'] == type_to_delete) &
            (df['cellule'] == cellule_to_delete)
        ]
        
        if rows_to_delete.empty:
            messagebox.showwarning("Avertissement", "Clé non trouvée")
            return
        
        df = df.drop(rows_to_delete.index)
        
        #vérifie et supprime la clé de la représentation de l'élément d'appartenance
        for _, row in rows_to_delete.iterrows():
            type_element = row["type"]
            if type_element in ["partie_mobile", "smalt", "boite_eclisse", "eclisse"]:
                self.delete_key_from_representation(book, type_element, appartenance_to_delete, nom_to_delete)
        
        book.remove(sheet)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        messagebox.showinfo("Information", f"Clé '{nom_to_delete}' avec appartenance '{appartenance_to_delete}' supprimée avec succès")
        
        #réinitialise les champs
        self.key_etat_var.set("prisonniere")
        self.penne_var.set("rentre")
        self.key_entries["appartenance"].delete(0, tk.END)
        self.key_entries["cellule"].delete(0, tk.END)
        self.key_entries["nom"].delete(0, tk.END)
        self.selected_row_index = None

    #supprime la clef de la représentation de son élément électrique d'appartenance si celui-ci est un smalt ou une partie mobile
    def delete_key_from_representation(self, book, type_element, appartenance, clef_nom):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"

        sheet_name = f"{type_element}"
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            return  
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        #trouve la ligne de représentation correspondant à l'appartenance
        row_index = df.index[df['nom'] == appartenance].tolist()
        if row_index:
            row = df.iloc[row_index[0] - 1]
            
            representation_list = ast.literal_eval(row['representation'])
            for i in range(len(representation_list)):
                for j in range(len(representation_list[i])):
                    if representation_list[i][j] == clef_nom:
                        representation_list[i][j] = 0
            
            new_representation = representation_list
            df.at[row_index[0], 'representation'] = str(new_representation)        
        
        book.remove(sheet)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"Clé '{clef_nom}' supprimée de la représentation de '{appartenance}' dans '{type_element}'")

    #met à jour la représentation de l'élément électrique d'appartenance suite à la modification d'une clef        
    def update_representation(self, book, type_element, appartenance):

        #la représentation sera actualisée en fonction du nombre de clefs
        if type_element == 'partie_mobile' or type_element == 'boite_eclisse':
            representation = [[0], [0]]
            positions_rentre = []
            positions_sorti = []
        elif type_element == 'smalt':
            representation = [ [0, 0]]
            positions_rentre = []
            positions_sorti = []
        elif type_element == 'eclisse':
            representation = [[1, 2, 1, 0], [0, 0, 0, 0]]
            positions_rentre = [(1, 2)]
            positions_sorti = [(1, 1)]
        else:
            return
        
        position_rentre_index = 0
        position_sorti_index = 0
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Feuille Excel 'clef' non trouvée")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        if type_element == 'partie_mobile':
            partie_mobile_position = self.partie_mobile_position_var.get()
            #on compte le nombre de clefs 'rentre' et 'sorti'
            rentre_count = len(df[(df["appartenance"] == appartenance) & (df["penne"] == "rentre")])
            sorti_count = len(df[(df["appartenance"] == appartenance) & (df["penne"] == "sorti")])
            print(rentre_count)
            if sorti_count >= rentre_count:
                for i in range(sorti_count):
                    representation[0].append(2)
                    representation[1].append(0)
                    representation[0].append(1)
                    representation[1].append(0)
                    positions_sorti.append((1, 2*i+1))
                    positions_rentre.append((1, 2*(i+1)))
                if partie_mobile_position == "debroche" and sorti_count == rentre_count:
                    representation[0].append(2)
                    representation[1].append(0)
                elif partie_mobile_position == "embroche":
                    representation[0][0] = 1
                    representation[0].append(0)
                    representation[1].append(0)
            else:
                for i in range(rentre_count-1):
                    representation[0].append(1)
                    representation[1].append(0)
                    representation[0].append(2)
                    representation[1].append(0)
                    positions_rentre.append((1, 2*i+1))
                    positions_sorti.append((1, 2*(i+1)))
                representation[0].append(1)
                representation[1].append(0)
                positions_rentre.append((1, 2*(rentre_count-1) + 1))
                if partie_mobile_position == "embroche":
                    representation[0][0] = 2
                    representation[0].append(0)
                    representation[1].append(0)
                elif partie_mobile_position == "debroche":
                    representation[0].append(2)
                    representation[1].append(0)
        elif type_element == 'boite_eclisse':
            boite_eclisse_position = self.boite_eclisse_position_var.get()
            #on compte le nombre de clefs 'rentre' et 'sorti'
            rentre_count = len(df[(df["appartenance"] == appartenance) & (df["penne"] == "rentre")])
            sorti_count = len(df[(df["appartenance"] == appartenance) & (df["penne"] == "sorti")])
            print(rentre_count)
            if sorti_count >= rentre_count:
                for i in range(sorti_count):
                    representation[0].append(2)
                    representation[1].append(0)
                    representation[0].append(1)
                    representation[1].append(0)
                    positions_sorti.append((1, 2*i+1))
                    positions_rentre.append((1, 2*(i+1)))
                if boite_eclisse_position == "debroche" and sorti_count == rentre_count:
                    representation[0].append(2)
                    representation[1].append(0)
                elif boite_eclisse_position == "embroche":
                    representation[0][0] = 1
                    representation[0].append(0)
                    representation[1].append(0)
            else:
                for i in range(rentre_count-1):
                    representation[0].append(1)
                    representation[1].append(0)
                    representation[0].append(2)
                    representation[1].append(0)
                    positions_rentre.append((1, 2*i+1))
                    positions_sorti.append((1, 2*(i+1)))
                representation[0].append(1)
                representation[1].append(0)
                positions_rentre.append((1, 2*(rentre_count-1) + 1))
                if boite_eclisse_position == "embroche":
                    representation[0][0] = 2
                    representation[0].append(0)
                    representation[1].append(0)
                elif boite_eclisse_position == "debroche":
                    representation[0].append(2)
                    representation[1].append(0)
        elif type_element == 'smalt':
            #on compte le nombre de clefs 'rentre' et 'sorti'
            rentre_count = len(df[(df["appartenance"] == appartenance) & (df["penne"] == "rentre")])
            sorti_count = len(df[(df["appartenance"] == appartenance) & (df["penne"] == "sorti")])
            print(rentre_count)
            if sorti_count >= rentre_count:
                for i in range(sorti_count):
                    representation.append([2, 0])
                    representation.append([1, 0])
                    positions_sorti.append((2*i+1, 1))
                    positions_rentre.append((2*(i+1), 1))
                representation.append([2, 0])
            else:
                for i in range(rentre_count):
                    representation.append([1, 0])
                    representation.append([2, 0])
                    positions_rentre.append((2*i+1, 1))
                    positions_sorti.append((2*(i+1), 1))   
        
        for _, row in df.iterrows():
            nom_element = row["appartenance"]
            if nom_element == appartenance:
                clef_nom = row["nom"]
                penne = row["penne"]
                if clef_nom:
                    if penne == "rentre" and position_rentre_index < len(positions_rentre):
                        r, c = positions_rentre[position_rentre_index]
                        representation[r][c] = clef_nom
                        position_rentre_index += 1
                    elif penne == "sorti" and position_sorti_index < len(positions_sorti):
                        r, c = positions_sorti[position_sorti_index]
                        representation[r][c] = clef_nom
                        position_sorti_index += 1                      
                    else:
                        messagebox.showwarning("Avertissement", "Vous voulez ajouter trop de clefs")

        sheet_name = f"{type_element}"
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            return  
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        #trouve la ligne de représentation correspondant à l'appartenance
        row_index = df.index[df['nom'] == appartenance].tolist()
        if row_index:
            row = df.iloc[row_index[0] - 1]
            df.at[row_index[0], 'representation'] = representation
        
        else:
            return
        
        
        book.remove(sheet)
        with pd.ExcelWriter(f"dataLHT/{self.palier}_modifs_LHT.xlsx", engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(df)
        
        print(f"Clé '{clef_nom}' ajoutée à la représentation de '{appartenance}'")

    # Crée l'onglet permettant de modifier les coffrets
    def create_coffret_tab(self):
        self.coffret_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.coffret_tab, text="Coffret")
        
        self.coffret_labels = ["nom", "type", "cellule"]
        self.coffret_entries = {}
        
        # Récupérer les valeurs possibles pour "nom" et "cellule" à partir du fichier Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "coffret"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            
            noms_existants = df['nom'].unique().tolist()  # Liste des noms existants
            cellules_existantes = df['cellule'].unique().tolist()  # Liste des cellules existantes
        else:
            noms_existants = []
            cellules_existantes = []

        # Création des labels et des combobox pour chaque champ
        for i, label in enumerate(self.coffret_labels):
            ttk.Label(self.coffret_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
            
            if label == "type":
                self.type_coffret_var = tk.StringVar(value="Coffret")
                type_coffret_menu = ttk.OptionMenu(self.coffret_tab, self.type_coffret_var, "Coffret", "Coffret", "CC")
                type_coffret_menu.grid(row=i, column=1, padx=10, pady=5)            
            elif label == "nom":
                # Combobox pour "nom"
                self.coffret_nom_var = tk.StringVar()
                nom_combobox = ttk.Combobox(self.coffret_tab, textvariable=self.coffret_nom_var, values=noms_existants)
                nom_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.coffret_entries[label] = nom_combobox
                nom_combobox.bind("<<ComboboxSelected>>", self.check_coffret_existence)
            elif label == "cellule":
                # Combobox pour "cellule"
                self.coffret_cellule_var = tk.StringVar()
                cellule_combobox = ttk.Combobox(self.coffret_tab, textvariable=self.coffret_cellule_var, values=cellules_existantes)
                cellule_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.coffret_entries[label] = cellule_combobox

        self.save_coffret_button = ttk.Button(self.coffret_tab, text="Sauvegarder", command=self.save_coffret)
        self.save_coffret_button.grid(row=len(self.coffret_labels), columnspan=2, padx=10, pady=10)
        
        self.delete_coffret_button = ttk.Button(self.coffret_tab, text="Supprimer", command=self.delete_coffret, state=tk.DISABLED)
        self.delete_coffret_button.grid(row=len(self.coffret_labels), column=3, padx=10, pady=10)

    # Vérifie si un coffret existe avec ce nom
    def check_coffret_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "coffret"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.coffret_entries["nom"].get()
        self.coffret_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.coffret_row_index:
            # Charge le coffret existant
            row = df.iloc[self.coffret_row_index[0] - 1]
            self.coffret_entries["cellule"].set(row["cellule"])
            self.delete_coffret_button.config(state=tk.NORMAL)

    # Vérifie si un coffret existe avec ce nom
    def check_coffret_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "coffret"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]

        nom_to_load = self.coffret_entries["nom"].get()
        self.coffret_row_index = df.index[df['nom'] == nom_to_load].tolist()

        # Charge le coffret existant si trouvé
        if self.coffret_row_index:
            row = df.iloc[self.coffret_row_index[0] - 1]
            self.coffret_entries["cellule"].set(row["cellule"])  # Remplir la cellule dans le combobox
            self.delete_coffret_button.config(state=tk.NORMAL)

    #supprime un coffret       
    def delete_coffret(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "coffret"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.coffret_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()

            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #Supprime les clefs qui ont comme appartenance le nom du coffret
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()

            print(f"Coffret supprimé dans {self.palier}_modifs.xlsx, feuille 'cellule'")
            
            for label in self.coffret_labels:
                self.coffret_entries[label].delete(0, tk.END)

    #sauvegarde les modifications 
    #si un coffret existe avec ce nom, il le modifie, sinon il crée un nouveau coffret    
    def save_coffret(self):
        new_coffret = {label: self.coffret_entries[label].get() for label in self.coffret_labels if label != "type"}
        new_coffret["type"] = "coffret"
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "coffret"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]

        if self.coffret_row_index:
            old_nom = df.loc[self.coffret_row_index[0], 'nom']
            # Modifie le coffret existant
            df.loc[self.coffret_row_index[0]] = new_coffret
            print(f"Modification coffret dans {self.palier}_modifs_LHT.xlsx, feuille 'coffret'")
            
            # Met à jour les clefs si le nom change
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_coffret['nom']
                
                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom du coffret")
            
            self.coffret_row_index = []
        else:
            df = df.append(new_coffret, ignore_index=True)
            print(f"Nouveaux coffrets sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'coffret'")

        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()

        print(f"Nouveaux coffrets sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'coffret'")

    '''
    #crée l'onglet permettant de modifier les fusibles   
    def create_fusible_tab(self):
        self.fusible_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.fusible_tab, text="Fusibles")
        
        self.fusible_labels = ["nom", "type", "cellule"]
        self.fusible_entries = {}
        
        for i, label in enumerate(self.fusible_labels):
            ttk.Label(self.fusible_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
            if label == "type":
                ttk.Label(self.fusible_tab, text="fusible").grid(row=i, column=1, padx=10, pady=5)
            else:
                entry = ttk.Entry(self.fusible_tab)
                entry.grid(row=i, column=1, padx=10, pady=5)
                self.fusible_entries[label] = entry
                if label == "nom":
                    entry.bind("<KeyRelease>", self.check_fusible_existence)
        
        self.save_fusible_button = ttk.Button(self.fusible_tab, text="Sauvegarder", command=self.save_fusibles)
        self.save_fusible_button.grid(row=len(self.fusible_labels), columnspan=2, padx=10, pady=10)
        self.delete_fusible_button = ttk.Button(self.fusible_tab, text="Supprimer", command=self.delete_fusible, state=tk.DISABLED)
        self.delete_fusible_button.grid(row=len(self.fusible_labels), column=3, padx=10, pady=10)
        
    #vérifie si un fusible avec ce nom existe, et le charge dans le cas échéant
    def check_fusible_existence(self, event=None):
        excel_path = f"{self.palier}_modifs_LHT.xlsx"
        sheet_name = "fusible"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.fusible_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if row_index:
            #charge le fusible existant
            row = df.iloc[row_index[0] - 1]
            self.fusible_entries["cellule"].delete(0, tk.END)
            self.fusible_entries["cellule"].insert(0, row["cellule"])
            self.delete_fusible_button.config(state=tk.NORMAL)

    #supprime un fusible existant     
    def delete_fusible(self):
        excel_path = f"{self.palier}_modifs_LHT.xlsx"
        sheet_name = "fusible"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.fusible_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            #supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            print(f"Cellule supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            for label in self.fusible_labels:
                self.fusible_entries[label].delete(0, tk.END)

    #sauvegarde les modifications 
    #si un fusible existe avec ce nom, il le modifie, sinon il crée un nouveau fusible  
    def save_fusibles(self):
        new_fusible = {label: self.fusible_entries[label].get() for label in self.fusible_labels if label != "type"}
        new_fusible["type"] = "fusible"
        
        excel_path = f"{self.palier}_modifs_LHT.xlsx"
        sheet_name = "fusible"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")

        
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        df = df.append(new_fusible, ignore_index=True)
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouveaux fusibles sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'fusible'")
    '''

    # Crée l'onglet permettant de modifier un panneau
    def create_panneau_tab(self):
        self.panneau_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.panneau_tab, text="Panneau")
        
        self.panneau_labels = ["nom", "type", "etat", "cellule"]
        self.panneau_entries = {}
        
        # Récupérer les valeurs possibles pour "nom" et "cellule" à partir du fichier Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "panneau"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            
            noms_existants = df['nom'].unique().tolist()  # Liste des noms existants
            cellules_existantes = df['cellule'].unique().tolist()  # Liste des cellules existantes
        else:
            noms_existants = []
            cellules_existantes = []

        # Création des labels et des combobox pour chaque champ
        for i, label in enumerate(self.panneau_labels):
            ttk.Label(self.panneau_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
            
            if label == "type":
                ttk.Label(self.panneau_tab, text="panneau").grid(row=i, column=1, padx=10, pady=5)
            elif label == "etat":
                self.panneau_etat_var = tk.StringVar(value="ferme")
                etat_menu = ttk.OptionMenu(self.panneau_tab, self.panneau_etat_var, "ferme", "ferme", "ouvert")
                etat_menu.grid(row=i, column=1, padx=10, pady=5)
            elif label == "nom":
                # Combobox pour "nom"
                self.panneau_nom_var = tk.StringVar()
                nom_combobox = ttk.Combobox(self.panneau_tab, textvariable=self.panneau_nom_var, values=noms_existants)
                nom_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.panneau_entries[label] = nom_combobox
                nom_combobox.bind("<<ComboboxSelected>>", self.check_panneau_existence)
            elif label == "cellule":
                # Combobox pour "cellule"
                self.panneau_cellule_var = tk.StringVar()
                cellule_combobox = ttk.Combobox(self.panneau_tab, textvariable=self.panneau_cellule_var, values=cellules_existantes)
                cellule_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.panneau_entries[label] = cellule_combobox

        self.save_panneau_button = ttk.Button(self.panneau_tab, text="Sauvegarder", command=self.save_panneau)
        self.save_panneau_button.grid(row=len(self.panneau_labels), columnspan=2, padx=10, pady=10)
        
        self.delete_panneau_button = ttk.Button(self.panneau_tab, text="Supprimer", command=self.delete_panneau, state=tk.DISABLED)
        self.delete_panneau_button.grid(row=len(self.panneau_labels), column=3, padx=10, pady=10)

    # Vérifie si un panneau existe avec ce nom, et le charge dans ce cas
    def check_panneau_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "panneau"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.panneau_entries["nom"].get()
        self.panneau_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.panneau_row_index:
            # Charge le panneau existant
            row = df.iloc[self.panneau_row_index[0] - 1]
            self.panneau_entries["cellule"].set(row["cellule"])
            self.delete_panneau_button.config(state=tk.NORMAL)

    #supprime un panneau existant       
    def delete_panneau(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "panneau"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.panneau_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                # Supprime les clefs qui ont comme appartenance le nom du panneau
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()

            print(f"Cellule supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            #Réinitialise les champs de saisie
            for label in self.panneau_labels:
                self.panneau_entries[label].delete(0, tk.END)
    
    #sauvegarde les modifications 
    #si un panneau existe avec ce nom, il le modifie, sinon il crée un nouveau panneau 
    def save_panneau(self):
        new_panneau = {label: self.panneau_entries[label].get() for label in self.panneau_labels if label not in ["type", "etat"]}
        new_panneau["type"] = "panneau"
        new_panneau["etat"] = self.panneau_etat_var.get()
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "panneau"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
        
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.panneau_row_index:
            old_nom = df.loc[self.panneau_row_index[0], 'nom']
            # Modifie le panneau existant
            df.loc[self.panneau_row_index[0]] = new_panneau
            print(f"Modification panneau dans {self.palier}_modifs_LHT.xlsx, feuille 'panneau'")
            
            # Met à jour les clefs dans la feuille clef
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_panneau['nom']

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom du panneau")
            else:
                messagebox.showwarning("Avertissement", "Feuille 'clef' non trouvée")
            self.panneau_row_index = []
        else:
            df = df.append(new_panneau, ignore_index=True)
            print(f"Nouveaux panneaux sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'panneau'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()        

    #crée l'onglet permettant de modifier les transformateurs
    def create_transformateur_tab(self):
        self.transformateur_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.transformateur_tab, text="Transformateur")
        
        self.transformateur_labels = ["nom", "type", "cellule"]
        self.transformateur_entries = {}
        
        # Récupérer les valeurs possibles pour "nom" et "cellule" à partir du fichier Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "transformateur"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            
            noms_existants = df['nom'].unique().tolist()  # Liste des noms existants
            cellules_existantes = df['cellule'].unique().tolist()  # Liste des cellules existantes
        else:
            noms_existants = []
            cellules_existantes = []

        # Création des labels et des combobox pour chaque champ
        for i, label in enumerate(self.transformateur_labels):
            ttk.Label(self.transformateur_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
            
            if label == "type":
                ttk.Label(self.transformateur_tab, text="transformateur").grid(row=i, column=1, padx=10, pady=5)
            elif label == "nom":
                # Combobox pour "nom"
                self.transformateur_nom_var = tk.StringVar()
                nom_combobox = ttk.Combobox(self.transformateur_tab, textvariable=self.transformateur_nom_var, values=noms_existants)
                nom_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.transformateur_entries[label] = nom_combobox
                nom_combobox.bind("<<ComboboxSelected>>", self.check_transformateur_existence)
            elif label == "cellule":
                # Combobox pour "cellule"
                self.transformateur_cellule_var = tk.StringVar()
                cellule_combobox = ttk.Combobox(self.transformateur_tab, textvariable=self.transformateur_cellule_var, values=cellules_existantes)
                cellule_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.transformateur_entries[label] = cellule_combobox

        self.save_transformateur_button = ttk.Button(self.transformateur_tab, text="Sauvegarder", command=self.save_transformateur)
        self.save_transformateur_button.grid(row=len(self.transformateur_labels), columnspan=2, padx=10, pady=10)
        
        self.delete_transformateur_button = ttk.Button(self.transformateur_tab, text="Supprimer", command=self.delete_transformateur, state=tk.DISABLED)
        self.delete_transformateur_button.grid(row=len(self.transformateur_labels), column=3, padx=10, pady=10)

    #vérifie si un transformateur existe avec ce nom, et le charge le cas échéant
    def check_transformateur_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "transformateur"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.transformateur_entries["nom"].get()
        self.transformateur_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.transformateur_row_index:
            # Charge le transformateur existant
            row = df.iloc[self.transformateur_row_index[0] - 1]
            self.transformateur_entries["cellule"].set(row["cellule"])
            self.delete_transformateur_button.config(state=tk.NORMAL)

    #supprime un transformateur existant        
    def delete_transformateur(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "transformateur"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.transformateur_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()

            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #Supprime les clefs qui ont comme appartenance le nom du transformateur
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
            
            print(f"Transformateur supprimé dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            for label in self.transformateur_labels:
                self.transformateur_entries[label].delete(0, tk.END)

    #sauvegarde les modifications 
    #si un transformateur existe avec ce nom, il le modifie, sinon il crée un nouveau transformateur     
    def save_transformateur(self):
        new_transformateur = {label: self.transformateur_entries[label].get() for label in self.transformateur_labels if label != "type"}
        new_transformateur["type"] = "transformateur"
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "transformateur"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.transformateur_row_index:
            old_nom = df.loc[self.transformateur_row_index[0], 'nom']
            df.loc[self.transformateur_row_index[0]] = new_transformateur
            print(f"Modification transformateur dans {self.palier}_modifs_LHT.xlsx, feuille 'transformateur'")
            
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_transformateur['nom']
                
                book.remove(clef_sheet)
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                    
            self.transformateur_row_index = []
        else:
            df = df.append(new_transformateur, ignore_index=True)
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()

        print(f"Nouveaux transformateurs sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'transformateur'")

    #crée l'onglet permettant de modifier les serrures mères
    def create_serrure_tab(self):
        self.serrure_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.serrure_tab, text="Serrure mère")
        
        self.serrure_labels = ["nom", "type", "cellule"]
        self.serrure_entries = {}

        # Remplir les combobox avec les valeurs possibles depuis le fichier Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "serrure"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            
            # Liste des valeurs pour les combobox
            noms_possibles = df['nom'].tolist()
            cellules_possibles = df['cellule'].tolist()

            for i, label in enumerate(self.serrure_labels):
                ttk.Label(self.serrure_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
                if label == "type":
                    self.type_serrure_var = tk.StringVar(value="serrure")
                    type_serrure_menu = ttk.OptionMenu(self.serrure_tab, self.type_serrure_var, "serrure", "serrure", "armoire")
                    type_serrure_menu.grid(row=i, column=1, padx=10, pady=5)
                elif label == "nom":
                    # Combobox pour 'nom' avec les valeurs disponibles dans le fichier Excel
                    nom_combobox = ttk.Combobox(self.serrure_tab, values=noms_possibles)
                    nom_combobox.grid(row=i, column=1, padx=10, pady=5)
                    self.serrure_entries[label] = nom_combobox
                    nom_combobox.bind("<<ComboboxSelected>>", self.check_serrure_existence)
                elif label == "cellule":
                    # Combobox pour 'cellule' avec les valeurs disponibles dans le fichier Excel
                    cellule_combobox = ttk.Combobox(self.serrure_tab, values=cellules_possibles)
                    cellule_combobox.grid(row=i, column=1, padx=10, pady=5)
                    self.serrure_entries[label] = cellule_combobox

        self.save_serrure_button = ttk.Button(self.serrure_tab, text="Sauvegarder", command=self.save_serrures)
        self.save_serrure_button.grid(row=len(self.serrure_labels), columnspan=2, padx=10, pady=10)
        self.delete_serrure_button = ttk.Button(self.serrure_tab, text="Supprimer", command=self.delete_serrure, state=tk.DISABLED)
        self.delete_serrure_button.grid(row=len(self.serrure_labels), column=3, padx=10, pady=10)

        # Liste les noms des serrures actuellement dans l'excel
        self.serrure_row_index = []

    #vérifie si une serrure mère existe avec ce nom, et la charge le cas échéant     
    def check_serrure_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "serrure"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.serrure_entries["nom"].get()
        self.serrure_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.serrure_row_index:
            # Charge la serrure existante
            row = df.iloc[self.serrure_row_index[0]]
            self.serrure_entries["cellule"].set(row["cellule"])  # Met à jour la cellule dans la combobox
            self.delete_serrure_button.config(state=tk.NORMAL)

    #supprime une serrure existante        
    def delete_serrure(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "serrure"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.serrure_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #supprime les clefs qui ont comme appartenance le nom de la serrure mère
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()

            print(f"Serrure supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            #Réinitialise les champs de saisie
            for label in self.serrure_labels:
                if label != 'type':
                    self.serrure_entries[label].delete(0, tk.END)

    #sauvegarde les modifications 
    #si une serrure mère existe avec ce nom, il la modifie, sinon il crée une nouvelle serrure     
    def save_serrures(self):
        new_serrure = {label: self.serrure_entries[label].get() for label in self.serrure_labels if label != "type"}
        new_serrure["type"] = "serrure_mere"
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "serrure"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.serrure_row_index:
            old_nom = df.loc[self.serrure_row_index[0], 'nom']
            # Modifie la serrure existante
            df.loc[self.serrure_row_index[0]] = new_serrure
            print(f"Modification serrure dans {self.palier}_modifs_LHT.xlsx, feuille 'serrure'")
            
            # Met à jour les clefs dans la feuille clef
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_serrure['nom']

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom de la serrure")
            else:
                messagebox.showwarning("Avertissement", "Feuille 'clef' non trouvée")
            self.serrure_row_index = []
        else:
            df = df.append(new_serrure, ignore_index=True)
            print(f"Nouvelles serrures sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'serrure'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouvelles serrures sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'serrure'")

    #crée l'onglet permettant de modifier les sources
    def create_source_tab(self):
        self.source_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.source_tab, text="Source")
        
        self.source_labels = ["nom", "type", "cellule"]
        self.source_entries = {}
        
        # Récupérer les valeurs possibles pour "nom" et "cellule" à partir du fichier Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "source"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            
            noms_existants = df['nom'].unique().tolist()  # Liste des noms existants
            cellules_existantes = df['cellule'].unique().tolist()  # Liste des cellules existantes
        else:
            noms_existants = []
            cellules_existantes = []

        # Création des labels et des combobox pour chaque champ
        for i, label in enumerate(self.source_labels):
            ttk.Label(self.source_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
            
            if label == "type":
                ttk.Label(self.source_tab, text="source").grid(row=i, column=1, padx=10, pady=5)
            elif label == "nom":
                # Combobox pour "nom"
                self.source_nom_var = tk.StringVar()
                nom_combobox = ttk.Combobox(self.source_tab, textvariable=self.source_nom_var, values=noms_existants)
                nom_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.source_entries[label] = nom_combobox
                nom_combobox.bind("<<ComboboxSelected>>", self.check_source_existence)
            elif label == "cellule":
                # Combobox pour "cellule"
                self.source_cellule_var = tk.StringVar()
                cellule_combobox = ttk.Combobox(self.source_tab, textvariable=self.source_cellule_var, values=cellules_existantes)
                cellule_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.source_entries[label] = cellule_combobox

        self.save_source_button = ttk.Button(self.source_tab, text="Sauvegarder", command=self.save_source)
        self.save_source_button.grid(row=len(self.source_labels), columnspan=2, padx=10, pady=10)
        
        self.delete_source_button = ttk.Button(self.source_tab, text="Supprimer", command=self.delete_source, state=tk.DISABLED)
        self.delete_source_button.grid(row=len(self.source_labels), column=3, padx=10, pady=10)

    #vérifie si un transformateur existe avec ce nom, et le charge le cas échéant
    def check_source_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "source"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.source_entries["nom"].get()
        self.source_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.source_row_index:
            # Charge le source existant
            row = df.iloc[self.source_row_index[0] - 1]
            self.source_entries["cellule"].set(row["cellule"])
            self.delete_source_button.config(state=tk.NORMAL)

    #supprime une source existante        
    def delete_source(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "source"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        nom_to_delete = self.source_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            # Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            print(f"Source supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            # Réinitialise les champs de saisie
            for label in self.source_labels:
                self.source_entries[label].set('')

    #sauvegarde les modifications 
    #si une source existe avec ce nom, il la modifie, sinon il crée une nouvelle source   
    def save_source(self):
        new_source = {label: self.source_entries[label].get() for label in self.source_labels if label != "type"}
        new_source["type"] = "source"
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "source"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = df.append(new_source, ignore_index=True)
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouvelles sources sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'source'")

    # Crée un onglet permettant de modifier les clefs libres
    def create_clefs_libres_tab(self):
        self.clefs_libres_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.clefs_libres_tab, text="Clés Libres")
        
        ttk.Label(self.clefs_libres_tab, text="clefs").grid(row=0, column=0, padx=10, pady=5)
        
        # Liste des valeurs possibles pour 'clefs'
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clefs_libres"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            possible_clefs = df['clefs'].tolist()  # Liste des clefs possibles
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        # Combobox pour 'clefs', rempli avec les valeurs existantes
        self.clefs_libres_combobox = ttk.Combobox(self.clefs_libres_tab, values=possible_clefs)
        self.clefs_libres_combobox.grid(row=0, column=1, padx=10, pady=5)
        ttk.Label(self.clefs_libres_tab, text="(sélectionner ou écrire une clef)").grid(row=0, column=2, padx=10, pady=5)
        
        self.save_clefs_libres_button = ttk.Button(self.clefs_libres_tab, text="Sauvegarder", command=self.save_clefs_libres)
        self.save_clefs_libres_button.grid(row=1, columnspan=2, padx=10, pady=10)
        
        self.delete_clefs_libres_button = ttk.Button(self.clefs_libres_tab, text="Supprimer", command=self.delete_clefs_libres)
        self.delete_clefs_libres_button.grid(row=1, column=2, padx=10, pady=10)

    # Supprime une clef existante   
    def delete_clefs_libres(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clefs_libres"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        clef_to_delete = self.clefs_libres_combobox.get()  # Utilise la combobox
        row_index = df.index[df['clefs'] == clef_to_delete].tolist()
        
        if row_index:
            # Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            print(f"Clef supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'clefs_libres'")
            
            # Réinitialise la combobox
            self.clefs_libres_combobox.set('')

    # Sauvegarde les modifications 
    # Si une clef libre existe avec ce nom, il la modifie, sinon il crée une nouvelle clef libre    
    def save_clefs_libres(self):
        new_clef_libre = {"clefs": self.clefs_libres_combobox.get()}  # Utilise la combobox
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clefs_libres"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        # Si la clef n'existe pas déjà, on l'ajoute
        if new_clef_libre["clefs"] not in df["clefs"].values:
            df = df.append(new_clef_libre, ignore_index=True)
            
            # Enregistre les modifications
            if sheet_name in book.sheetnames:
                book.remove(sheet)
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            print(f"Nouvelles clefs_libres sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'clefs_libres'")
        else:
            messagebox.showinfo("Info", "Cette clef existe déjà.")

    #crée l'onglet permettant de modifier une partie mobile
    def create_partie_mobile_tab(self):
        self.partie_mobile_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.partie_mobile_tab, text="Partie mobile")

        self.partie_mobile_labels = ["nom", "type", "cellule", "etat", "position", "representation", "deplacement"]
        self.partie_mobile_entries = {}

        # Charger les valeurs possibles depuis Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "partie_mobile"

        try:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                df = pd.DataFrame(sheet.values)
                df.columns = df.iloc[0]
                df = df[1:]
                noms_possibles = list(df["nom"].dropna().unique())
                cellules_possibles = list(df["cellule"].dropna().unique())
            else:
                noms_possibles, cellules_possibles = [], []
        except Exception as e:
            print("Erreur lors du chargement de l'Excel:", e)
            noms_possibles, cellules_possibles = [], []

        for i, label in enumerate(self.partie_mobile_labels):
            ttk.Label(self.partie_mobile_tab, text=label).grid(row=i, column=0, padx=10, pady=5)

            if label == "type":
                self.partie_mobile_type_var = tk.StringVar(value="contacteur")
                type_menu = ttk.OptionMenu(self.partie_mobile_tab, self.partie_mobile_type_var, "contacteur", "contacteur", "disjoncteur", "tiroir")
                type_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "etat":
                self.partie_mobile_etat_var = tk.StringVar(value="ouvert")
                etat_menu = ttk.OptionMenu(self.partie_mobile_tab, self.partie_mobile_etat_var, "ouvert", "ouvert", "ferme")
                etat_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "position":
                self.partie_mobile_position_var = tk.StringVar(value="embroche")
                position_menu = ttk.OptionMenu(self.partie_mobile_tab, self.partie_mobile_position_var, "embroche", "embroche", "debroche", "bloque")
                position_menu.grid(row=i, column=1, padx=10, pady=5)
                self.partie_mobile_position_var.trace_add("write", self.combined_keyrelease_handler_partie_mobile)

            elif label == "deplacement":
                self.partie_mobile_deplacement_var = tk.StringVar(value="haut")
                deplacement_menu = ttk.OptionMenu(self.partie_mobile_tab, self.partie_mobile_deplacement_var, "haut", "haut", "bas")
                deplacement_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "nom":
                self.partie_mobile_nom_var = tk.StringVar()
                self.partie_mobile_entries[label] = ttk.Combobox(self.partie_mobile_tab, textvariable=self.partie_mobile_nom_var, values=noms_possibles)
                self.partie_mobile_entries[label].grid(row=i, column=1, padx=10, pady=5)
                self.partie_mobile_entries[label].bind("<<ComboboxSelected>>", self.combined_keyrelease_handler_partie_mobile)

            elif label == "cellule":
                self.partie_mobile_cellule_var = tk.StringVar()
                self.partie_mobile_entries[label] = ttk.Combobox(self.partie_mobile_tab, textvariable=self.partie_mobile_cellule_var, values=cellules_possibles)
                self.partie_mobile_entries[label].grid(row=i, column=1, padx=10, pady=5)

            else:
                entry = ttk.Entry(self.partie_mobile_tab)
                entry.grid(row=i, column=1, padx=10, pady=5)
                self.partie_mobile_entries[label] = entry
                if label == "representation":
                    ttk.Label(self.partie_mobile_tab, text="(se remplit automatiquement, ne pas modifier sauf erreur)").grid(row=i, column=2, padx=10, pady=5)

        self.save_partie_mobile_button = ttk.Button(self.partie_mobile_tab, text="Sauvegarder", command=self.save_partie_mobile)
        self.save_partie_mobile_button.grid(row=len(self.partie_mobile_labels), columnspan=2, padx=10, pady=10)

        self.delete_partie_mobile_button = ttk.Button(self.partie_mobile_tab, text="Supprimer", command=self.delete_partie_mobile, state=tk.DISABLED)
        self.delete_partie_mobile_button.grid(row=len(self.partie_mobile_labels), columnspan=3, padx=10, pady=10)

    #vérifie si partie mobile existe avec ce nom, et la charge le cas échéant
    def check_partie_mobile_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "partie_mobile"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.partie_mobile_entries["nom"].get()
        self.partie_mobile_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.partie_mobile_row_index:
            #Charge la partie mobile existant
            row = df.iloc[self.partie_mobile_row_index[0] - 1]
            self.partie_mobile_entries["cellule"].delete(0, tk.END)
            self.partie_mobile_entries["cellule"].insert(0, row["cellule"])
            self.partie_mobile_entries["representation"].delete(0, tk.END)
            self.partie_mobile_entries["representation"].insert(0, row["representation"])
            self.partie_mobile_etat_var.set(row["etat"])
            self.partie_mobile_type_var.set(row["type"])
            self.partie_mobile_position_var.set(row["position"])
            self.partie_mobile_deplacement_var.set(row["deplacement"])
            self.delete_partie_mobile_button.config(state=tk.NORMAL)
 
    def combined_keyrelease_handler_partie_mobile(self, *args, event=None):
        self.update_representation_partie_mobile()
        self.check_partie_mobile_existence(event)
               
    #sauvegarde les modifications 
    #si une partie mobile existe avec ce nom, il la modifie, sinon il crée une nouvelle partie mobile
    def save_partie_mobile(self):
        new_partie_mobile = {label: self.partie_mobile_entries[label].get() for label in self.partie_mobile_labels if label not in ["type", "etat", "position", "deplacement"]}
        new_partie_mobile["type"] = self.partie_mobile_type_var.get()
        new_partie_mobile["etat"] = "ouvert"
        new_partie_mobile["position"] = self.partie_mobile_position_var.get()
        new_partie_mobile["deplacement"] = self.partie_mobile_deplacement_var.get()
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "partie_mobile"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
    
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.partie_mobile_row_index:
            old_nom = df.loc[self.partie_mobile_row_index[0], 'nom']
            #Modifie la partie_mobile existante
            df.loc[self.partie_mobile_row_index[0]] = new_partie_mobile
            print(f"Modification partie_mobile dans {self.palier}_modifs_LHT.xlsx, feuille 'partie_mobile'")
            #Met à jour les clefs dans la feuille clef
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_partie_mobile['nom']

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom du partie_mobile")
            else:
                messagebox.showwarning("Avertissement", "Feuille 'clef' non trouvée")
            self.partie_mobile_row_index = []
        else:
            df = df.append(new_partie_mobile, ignore_index=True)
            print(f"Nouveaux partie_mobilex sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'partie_mobile'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouvelles parties mobiles sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'partie_mobile'")

    #supprime une partie mobile existante    
    def delete_partie_mobile(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "partie_mobile"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.partie_mobile_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()

            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #Supprime les clefs qui ont comme appartenance le nom de la partie mobile
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
            
            print(f"Partie mobile supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            #Réinitialise les champs de saisie
            for label in self.partie_mobile_labels:
                self.partie_mobile_entries[label].delete(0, tk.END)

    #met à jour la représentation de la partie mobile
    def update_representation_partie_mobile(self, event=None):
        representation = [[0], [0]]
        positions_rentre = []
        positions_sorti = []
        
        partie_mobile_nom = self.partie_mobile_nom_var.get()
        partie_mobile_position = self.partie_mobile_position_var.get()
        position_rentre_index = 0
        position_sorti_index = 0
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Feuille Excel 'clef' non trouvée")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]

        #compte le nombre de clefs 'rentre' et 'sorti'
        rentre_count = len(df[(df["appartenance"] == partie_mobile_nom) & (df["penne"] == "rentre")])
        sorti_count = len(df[(df["appartenance"] == partie_mobile_nom) & (df["penne"] == "sorti")])
        print(rentre_count)
        if sorti_count >= rentre_count:
            for i in range(sorti_count):
                representation[0].append(2)
                representation[1].append(0)
                representation[0].append(1)
                representation[1].append(0)
                positions_sorti.append((1, 2*i+1))
                positions_rentre.append((1, 2*(i+1)))
            if partie_mobile_position == "debroche" and sorti_count == rentre_count:
                representation[0].append(2)
                representation[1].append(0)
            elif partie_mobile_position == "embroche":
                representation[0][0] = 1
                representation[0].append(0)
                representation[1].append(0)
        else:
            for i in range(rentre_count-1):
                representation[0].append(1)
                representation[1].append(0)
                representation[0].append(2)
                representation[1].append(0)
                positions_rentre.append((1, 2*i+1))
                positions_sorti.append((1, 2*(i+1)))
            representation[0].append(1)
            representation[1].append(0)
            positions_rentre.append((1, 2*(rentre_count-1) + 1))
            if partie_mobile_position == "embroche":
                representation[0][0] = 2
                representation[0].append(0)
                representation[1].append(0)
            elif partie_mobile_position == "debroche":
                representation[0].append(2)
                representation[1].append(0)
        print(representation)
        print(positions_rentre)
        print(positions_sorti)

        #on parcourt toutes les clés dans la feuille Excel
        for _, row in df.iterrows():
            appartenance = row["appartenance"]
            if appartenance == partie_mobile_nom:
                clef_nom = row["nom"]
                penne = row["penne"]
                cellule = row["cellule"]
                if clef_nom:
                    if penne == "rentre" and position_rentre_index < len(positions_rentre):
                        r, c = positions_rentre[position_rentre_index]
                        representation[r][c] = clef_nom
                        position_rentre_index += 1
                    elif penne == "sorti" and position_sorti_index < len(positions_sorti):
                        r, c = positions_sorti[position_sorti_index]
                        representation[r][c] = clef_nom
                        position_sorti_index += 1                      
                       
                    #on met à jour la case "cellule" avec la valeur de la clé
                    self.partie_mobile_entries["cellule"].delete(0, tk.END)
                    self.partie_mobile_entries["cellule"].insert(0, cellule)
                    
        
        #on met à jour le champ de saisie de la représentation
        self.partie_mobile_entries["representation"].delete(0, tk.END)
        self.partie_mobile_entries["representation"].insert(0, str(representation))
        
    #crée l'onglet permettant de modifier les éclisses
    def create_eclisse_tab(self):
        self.eclisse_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.eclisse_tab, text="Eclisse")

        self.eclisse_labels = ["nom", "type", "cellule", "etat", "position", "representation", "deplacement", "voie"]
        self.eclisse_entries = {}

        # Charger les valeurs possibles depuis Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "eclisse"

        try:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                df = pd.DataFrame(sheet.values)
                df.columns = df.iloc[0]
                df = df[1:]
                noms_possibles = list(df["nom"].dropna().unique())
                cellules_possibles = list(df["cellule"].dropna().unique())
            else:
                noms_possibles, cellules_possibles = [], []
        except Exception as e:
            print("Erreur lors du chargement de l'Excel:", e)
            noms_possibles, cellules_possibles = [], []

        for i, label in enumerate(self.eclisse_labels):
            ttk.Label(self.eclisse_tab, text=label).grid(row=i, column=0, padx=10, pady=5)

            if label == "type":
                self.eclisse_type_var = tk.StringVar(value="eclisse")
                ttk.Label(self.eclisse_tab, textvariable=self.eclisse_type_var).grid(row=i, column=1, padx=10, pady=5)

            elif label == "etat":
                self.eclisse_etat_var = tk.StringVar(value="fil")
                ttk.Label(self.eclisse_tab, textvariable=self.eclisse_etat_var).grid(row=i, column=1, padx=10, pady=5)

            elif label == "position":
                self.eclisse_position_var = tk.StringVar(value="presente")
                position_menu = ttk.OptionMenu(self.eclisse_tab, self.eclisse_position_var, "presente", "presente", "absente")
                position_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "voie":
                self.eclisse_voie_var = tk.StringVar(value="A")
                voie_menu = ttk.OptionMenu(self.eclisse_tab, self.eclisse_voie_var, "A", "A", "B")
                voie_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "deplacement":
                self.eclisse_deplacement_var = tk.StringVar(value="aucun")
                ttk.Label(self.eclisse_tab, textvariable=self.eclisse_deplacement_var).grid(row=i, column=1, padx=10, pady=5)

            elif label == "nom":
                self.eclisse_nom_var = tk.StringVar()
                self.eclisse_entries[label] = ttk.Combobox(self.eclisse_tab, textvariable=self.eclisse_nom_var, values=noms_possibles)
                self.eclisse_entries[label].grid(row=i, column=1, padx=10, pady=5)
                self.eclisse_entries[label].bind("<<ComboboxSelected>>", self.combined_keyrelease_handler_eclisse)

            elif label == "cellule":
                self.eclisse_cellule_var = tk.StringVar()
                self.eclisse_entries[label] = ttk.Combobox(self.eclisse_tab, textvariable=self.eclisse_cellule_var, values=cellules_possibles)
                self.eclisse_entries[label].grid(row=i, column=1, padx=10, pady=5)
            else:
                entry = ttk.Entry(self.eclisse_tab)
                entry.grid(row=i, column=1, padx=10, pady=5)
                self.eclisse_entries[label] = entry
                
                if label == "representation":
                    ttk.Label(self.eclisse_tab, text="(se remplit automatiquement, ne pas modifier)").grid(row=i, column=2, padx=10, pady=5)

        self.save_eclisse_button = ttk.Button(self.eclisse_tab, text="Sauvegarder", command=self.save_eclisse)
        self.save_eclisse_button.grid(row=len(self.eclisse_labels), columnspan=2, padx=10, pady=10)
        self.delete_eclisse_button = ttk.Button(self.eclisse_tab, text="Supprimer", command=self.delete_eclisse, state=tk.DISABLED)
        self.delete_eclisse_button.grid(row=len(self.eclisse_labels), columnspan=3, padx=10, pady=10)
        
    #vérifie si une éclisse existe avec ce nom, et la charge le cas échéant    
    def check_eclisse_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "eclisse"

        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return

        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]

        nom_to_load = self.eclisse_nom_var.get()
        self.eclisse_row_index = df.index[df['nom'] == nom_to_load].tolist()

        if self.eclisse_row_index:
            row = df.iloc[self.eclisse_row_index[0] - 1]
            self.eclisse_entries["cellule"].delete(0, tk.END)
            self.eclisse_entries["cellule"].insert(0, row["cellule"])
            self.eclisse_entries["representation"].delete(0, tk.END)
            self.eclisse_entries["representation"].insert(0, row["representation"])
            self.eclisse_type_var.set(row["type"])
            self.eclisse_etat_var.set(row["etat"])
            self.eclisse_position_var.set(row["position"])
            self.eclisse_deplacement_var.set(row["deplacement"])
            self.eclisse_voie_var.set(row["voie"])

            self.delete_eclisse_button.config(state=tk.NORMAL)
            
    def combined_keyrelease_handler_eclisse(self, event=None):
        self.update_representation_eclisse()
        self.check_eclisse_existence(event)

    #sauvegarde les modifications 
    #si une éclisse existe avec ce nom, il la modifie, sinon il crée une nouvelle éclisse
    def save_eclisse(self):
        new_eclisse = {label: self.eclisse_entries[label].get() for label in self.eclisse_labels if label not in ["type", "etat", "position", "deplacement", "voie"]}
        new_eclisse["type"] = self.eclisse_type_var.get()
        new_eclisse["etat"] = self.eclisse_etat_var.get()
        new_eclisse["position"] = self.eclisse_position_var.get()
        new_eclisse["deplacement"] = self.eclisse_deplacement_var.get()
        new_eclisse["voie"] = self.eclisse_voie_var.get()
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "eclisse"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
    
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.eclisse_row_index:
            old_nom = df.loc[self.eclisse_row_index[0], 'nom']
            #Modifie l'eclisse existante
            df.loc[self.eclisse_row_index[0]] = new_eclisse
            print(f"Modification eclisse dans {self.palier}_modifs.xlsx, feuille 'eclisse'")
            #Met à jour les clefs dans la feuille clef
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_eclisse['nom']

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom du eclisse")
            else:
                messagebox.showwarning("Avertissement", "Feuille 'clef' non trouvée")
            self.eclisse_row_index = []
        else:
            df = df.append(new_eclisse, ignore_index=True)
            print(f"Nouveaux eclisses sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'eclisse'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouvelles eclisses sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'eclisse'")

    #supprime une éclisse existante    
    def delete_eclisse(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "eclisse"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.eclisse_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()

            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #Supprime les clefs qui ont comme appartenance le nom de l'eclisse
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
            
            print(f"eclisse supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'eclisse'")
            
            #Réinitialise les champs de saisie
            for label in self.boite_eclisse_labels:
                if label in self.boite_eclisse_entries:
                    self.boite_eclisse_entries[label].delete(0, tk.END)

    #met à jour la représentation de la serrurerie en fonction des clefs qui appartiennent à l'éclisse 
    def update_representation_eclisse(self, event=None):
        representation = [[1, 2, 1, 0], [0, 0, 0, 0]]

        eclisse_nom = self.eclisse_nom_var.get()

        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"

        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Feuille Excel 'clef' non trouvée")
            return

        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]

        for _, row in df.iterrows():
            appartenance = row["appartenance"]
            if appartenance == eclisse_nom:
                self.eclisse_cellule_var.set(row["cellule"])                    
        
        #on met à jour le champ de saisie de la représentation
        self.eclisse_entries["representation"].delete(0, tk.END)
        self.eclisse_entries["representation"].insert(0, str(representation))

    #crée l'onglet permettant de modifier les boites à eclisses
    def create_boite_eclisse_tab(self):
        self.boite_eclisse_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.boite_eclisse_tab, text="Boite à éclisses")

        self.boite_eclisse_labels = ["nom", "type", "cellule", "position", "representation", "deplacement", "voie", "stock_eclisses"]
        self.boite_eclisse_entries = {}

        # Récupération des valeurs uniques pour "nom" et "cellule"
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "boite_eclisse"
        
        try:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                df = pd.DataFrame(sheet.values)
                df.columns = df.iloc[0]
                df = df[1:]
                noms_possibles = list(df["nom"].dropna().unique())
                cellules_possibles = list(df["cellule"].dropna().unique())
            else:
                noms_possibles, cellules_possibles = [], []
        except Exception as e:
            print("Erreur lors du chargement de l'Excel:", e)
            noms_possibles, cellules_possibles = [], []

        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            noms_possibles = df["nom"].dropna().unique().tolist()
            cellules_possibles = df["cellule"].dropna().unique().tolist()

        for i, label in enumerate(self.boite_eclisse_labels):
            ttk.Label(self.boite_eclisse_tab, text=label).grid(row=i, column=0, padx=10, pady=5)

            if label == "type":
                self.boite_eclisse_type_var = tk.StringVar(value="boite_eclisse")
                ttk.Label(self.boite_eclisse_tab, textvariable=self.boite_eclisse_type_var).grid(row=i, column=1, padx=10, pady=5)

            elif label == "position":
                self.boite_eclisse_position_var = tk.StringVar(value="embroche")
                position_menu = ttk.OptionMenu(self.boite_eclisse_tab, self.boite_eclisse_position_var, "embroche", "embroche", "debroche")
                position_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "voie":
                self.boite_eclisse_voie_var = tk.StringVar(value="A")
                voie_menu = ttk.OptionMenu(self.boite_eclisse_tab, self.boite_eclisse_voie_var, "A", "A", "B")
                voie_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "stock_eclisses":
                self.boite_eclisse_stock_var = tk.StringVar(value="0")
                stock_menu = ttk.OptionMenu(self.boite_eclisse_tab, self.boite_eclisse_stock_var, "0", "0", "1")
                stock_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "deplacement":
                self.boite_eclisse_deplacement_var = tk.StringVar(value="droite")
                deplacement_menu = ttk.OptionMenu(self.boite_eclisse_tab, self.boite_eclisse_deplacement_var, "droite", "droite", "gauche", "pas")
                deplacement_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "nom":
                self.boite_eclisse_nom_var = tk.StringVar()
                self.boite_eclisse_entries[label] = ttk.Combobox(self.boite_eclisse_tab, textvariable=self.boite_eclisse_nom_var, values=noms_possibles)
                self.boite_eclisse_entries[label].grid(row=i, column=1, padx=10, pady=5)
                self.boite_eclisse_entries[label].bind("<<ComboboxSelected>>", self.combined_keyrelease_handler_boite_eclisse)

            elif label == "cellule":
                self.boite_eclisse_cellule_var = tk.StringVar()
                self.boite_eclisse_entries[label] = ttk.Combobox(self.boite_eclisse_tab, textvariable=self.boite_eclisse_cellule_var, values=cellules_possibles)
                self.boite_eclisse_entries[label].grid(row=i, column=1, padx=10, pady=5)

            else:
                entry = ttk.Entry(self.boite_eclisse_tab)
                entry.grid(row=i, column=1, padx=10, pady=5)
                self.boite_eclisse_entries[label] = entry

                if label == "representation":
                    ttk.Label(self.boite_eclisse_tab, text="(se remplit automatiquement, ne pas modifier sauf si erreur)").grid(row=i, column=2, padx=10, pady=5)

        self.save_boite_eclisse_button = ttk.Button(self.boite_eclisse_tab, text="Sauvegarder", command=self.save_boite_eclisse)
        self.save_boite_eclisse_button.grid(row=len(self.boite_eclisse_labels), columnspan=2, padx=10, pady=10)
        self.delete_boite_eclisse_button = ttk.Button(self.boite_eclisse_tab, text="Supprimer", command=self.delete_boite_eclisse, state=tk.DISABLED)
        self.delete_boite_eclisse_button.grid(row=len(self.boite_eclisse_labels), columnspan=3, padx=10, pady=10)

    #vérifie si une boite à éclisses existe, et la charge le cas échéant   
    def check_boite_eclisse_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "boite_eclisse"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.boite_eclisse_entries["nom"].get()
        self.boite_eclisse_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.boite_eclisse_row_index:
            #Charge la boite à eclisses existante
            row = df.iloc[self.boite_eclisse_row_index[0] - 1]
            self.boite_eclisse_type_var.set(row["type"])
            self.boite_eclisse_entries["cellule"].delete(0, tk.END)
            self.boite_eclisse_entries["cellule"].insert(0, row["cellule"])
            self.boite_eclisse_entries["representation"].delete(0, tk.END)
            self.boite_eclisse_entries["representation"].insert(0, row["representation"])
            self.boite_eclisse_position_var.set(row["position"])
            self.boite_eclisse_deplacement_var.set(row["deplacement"])
            self.boite_eclisse_voie_var.set(row["voie"])
            self.boite_eclisse_stock_var.set(row["stock_eclisses"])

            self.delete_boite_eclisse_button.config(state=tk.NORMAL)
            
    def combined_keyrelease_handler_boite_eclisse(self, event=None):
        self.update_representation_boite_eclisse()
        self.check_boite_eclisse_existence(event)
               
    #sauvegarde les modifications 
    #si une boite à eclisse existe avec ce nom, il la modifie, sinon il crée une nouvelle boite à éclisses 
    def save_boite_eclisse(self):
        new_boite_eclisse = {label: self.boite_eclisse_entries[label].get() for label in self.boite_eclisse_labels if label not in ["type", "etat", "position", "deplacement", "voie", "stock_eclisses"]}
        new_boite_eclisse["type"] = self.boite_eclisse_type_var.get()
        new_boite_eclisse["position"] = self.boite_eclisse_position_var.get()
        new_boite_eclisse["deplacement"] = self.boite_eclisse_deplacement_var.get()
        new_boite_eclisse["voie"] = self.boite_eclisse_voie_var.get()
        new_boite_eclisse["stock_eclisses"] = int(self.boite_eclisse_stock_var.get())
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "boite_eclisse"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
    
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.boite_eclisse_row_index:
            old_nom = df.loc[self.boite_eclisse_row_index[0], 'nom']
            #Modifie la boite_eclisse existante
            df.loc[self.boite_eclisse_row_index[0]] = new_boite_eclisse
            print(f"Modification boite_eclisse dans {self.palier}_modifs_LHT.xlsx, feuille 'boite_eclisse'")
            #Met à jour les clefs dans la feuille clef
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_boite_eclisse['nom']

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom du boite_eclisse")
            else:
                messagebox.showwarning("Avertissement", "Feuille 'clef' non trouvée")
            self.boite_eclisse_row_index = []
        else:
            df = df.append(new_boite_eclisse, ignore_index=True)
            print(f"Nouveaux boite_eclissex sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'boite_eclisse'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouvelles boites à eclisses sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'boite_eclisse'")

    #supprime la boite à éclisses existante   
    def delete_boite_eclisse(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "boite_eclisse"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.boite_eclisse_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()

            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #Supprime les clefs qui ont comme appartenance le nom de la partie mobile
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
            
            print(f"Boite à eclisses supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'boite_eclisse'")
            
            #Réinitialise les champs de saisie
            for label in self.boite_eclisse_labels:
                if label in self.boite_eclisse_entries:
                    self.boite_eclisse_entries[label].delete(0, tk.END)

    def update_representation_boite_eclisse(self, event=None):
        representation = [[0], [0]]
        
        boite_eclisse_nom = self.boite_eclisse_entries.get("nom", "").get()
        boite_eclisse_position = self.boite_eclisse_position_var.get()
        
        positions_rentre = []
        positions_sorti = []
        
        position_rentre_index = 0
        position_sorti_index = 0
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Feuille Excel 'clef' non trouvée")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        #compte le nombre de clefs 'rentre' et 'sorti'
        rentre_count = len(df[(df["appartenance"] == boite_eclisse_nom) & (df["penne"] == "rentre")])
        sorti_count = len(df[(df["appartenance"] == boite_eclisse_nom) & (df["penne"] == "sorti")])
        print(rentre_count)
        if sorti_count >= rentre_count:
            for i in range(sorti_count):
                representation[0].append(2)
                representation[1].append(0)
                representation[0].append(1)
                representation[1].append(0)
                positions_sorti.append((1, 2*i+1))
                positions_rentre.append((1, 2*(i+1)))
            if boite_eclisse_position == "debroche" and sorti_count == rentre_count:
                representation[0].append(2)
                representation[1].append(0)
            elif boite_eclisse_position == "embroche":
                representation[0][0] = 1
                representation[0].append(0)
                representation[1].append(0)
        else:
            for i in range(rentre_count-1):
                representation[0].append(1)
                representation[1].append(0)
                representation[0].append(2)
                representation[1].append(0)
                positions_rentre.append((1, 2*i+1))
                positions_sorti.append((1, 2*(i+1)))
            representation[0].append(1)
            representation[1].append(0)
            positions_rentre.append((1, 2*(rentre_count-1) + 1))
            if boite_eclisse_position == "embroche":
                representation[0][0] = 2
                representation[0].append(0)
                representation[1].append(0)
            elif boite_eclisse_position == "debroche":
                representation[0].append(2)
                representation[1].append(0)
        print(representation)
        print(positions_rentre)
        print(positions_sorti)

        for _, row in df.iterrows():
            appartenance = row["appartenance"]
            if appartenance == boite_eclisse_nom:
                clef_nom = row["nom"]
                penne = row["penne"]
                cellule = row["cellule"]
                if clef_nom:
                    if penne == "rentre" and position_rentre_index < len(positions_rentre):
                        r, c = positions_rentre[position_rentre_index]
                        representation[r][c] = clef_nom
                        position_rentre_index += 1
                    elif penne == "sorti" and position_sorti_index < len(positions_sorti):
                        r, c = positions_sorti[position_sorti_index]
                        representation[r][c] = clef_nom
                        position_sorti_index += 1                      
                    else:
                        messagebox.showwarning("Avertissement", "Vous voulez ajouter trop de clefs")
                        
                    self.boite_eclisse_entries["cellule"].delete(0, tk.END)
                    self.boite_eclisse_entries["cellule"].insert(0, cellule)
                    
        
        #on met à jour le champ de saisie de la représentation
        self.boite_eclisse_entries["representation"].delete(0, tk.END)
        self.boite_eclisse_entries["representation"].insert(0, str(representation))

    #crée l'onglet permettant de modifier les smalts
    def create_smalt_tab(self):
        self.smalt_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.smalt_tab, text="Smalt")

        self.smalt_labels = ["nom", "type", "cellule", "etat", "representation", "deplacement"]
        self.smalt_entries = {}

        # Charger les valeurs possibles depuis Excel
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "smalt"

        try:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                df = pd.DataFrame(sheet.values)
                df.columns = df.iloc[0]
                df = df[1:]
                noms_possibles = list(df["nom"].dropna().unique())
                cellules_possibles = list(df["cellule"].dropna().unique())
            else:
                noms_possibles, cellules_possibles = [], []
        except Exception as e:
            print("Erreur lors du chargement de l'Excel:", e)
            noms_possibles, cellules_possibles = [], []

        for i, label in enumerate(self.smalt_labels):
            ttk.Label(self.smalt_tab, text=label).grid(row=i, column=0, padx=10, pady=5)

            if label == "type":
                self.smalt_type_var = tk.StringVar(value="smalt")
                ttk.Label(self.smalt_tab, textvariable=self.smalt_type_var).grid(row=i, column=1, padx=10, pady=5)

            elif label == "etat":
                self.smalt_etat_var = tk.StringVar(value="ouvert")
                etat_menu = ttk.OptionMenu(self.smalt_tab, self.smalt_etat_var, "ouvert", "ouvert", "ferme")
                etat_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "deplacement":
                self.smalt_deplacement_var = tk.StringVar(value="haut")
                deplacement_menu = ttk.OptionMenu(self.smalt_tab, self.smalt_deplacement_var, "haut", "haut", "bas")
                deplacement_menu.grid(row=i, column=1, padx=10, pady=5)

            elif label == "nom":
                self.smalt_nom_var = tk.StringVar()
                self.smalt_entries[label] = ttk.Combobox(self.smalt_tab, textvariable=self.smalt_nom_var, values=noms_possibles)
                self.smalt_entries[label].grid(row=i, column=1, padx=10, pady=5)
                self.smalt_entries[label].bind("<<ComboboxSelected>>", self.combined_keyrelease_handler_smalt)

            elif label == "cellule":
                self.smalt_cellule_var = tk.StringVar()
                self.smalt_entries[label] = ttk.Combobox(self.smalt_tab, textvariable=self.smalt_cellule_var, values=cellules_possibles)
                self.smalt_entries[label].grid(row=i, column=1, padx=10, pady=5)

            else:
                entry = ttk.Entry(self.smalt_tab)
                entry.grid(row=i, column=1, padx=10, pady=5)
                self.smalt_entries[label] = entry
                if label == "representation":
                    ttk.Label(self.smalt_tab, text="(se remplit automatiquement, ne pas modifier sauf erreur)").grid(row=i, column=2, padx=10, pady=5)

        self.save_smalt_button = ttk.Button(self.smalt_tab, text="Sauvegarder", command=self.save_smalt)
        self.save_smalt_button.grid(row=len(self.smalt_labels), columnspan=2, padx=10, pady=10)

        self.delete_smalt_button = ttk.Button(self.smalt_tab, text="Supprimer", command=self.delete_smalt, state=tk.DISABLED)
        self.delete_smalt_button.grid(row=len(self.smalt_labels), columnspan=3, padx=10, pady=10)

    #vérifie si un smalt existe avec ce nom, et le charge le cas échéant    
    def check_smalt_existence(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "smalt"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.smalt_entries["nom"].get()
        self.smalt_row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if self.smalt_row_index:
            #Charge le smalt existant
            row = df.iloc[self.smalt_row_index[0] - 1]
            self.smalt_entries["cellule"].delete(0, tk.END)
            self.smalt_entries["cellule"].insert(0, row["cellule"])
            self.smalt_entries["representation"].delete(0, tk.END)
            self.smalt_entries["representation"].insert(0, row["representation"])
            self.smalt_etat_var.set(row["etat"])
            self.smalt_deplacement_var.set(row["deplacement"])
            self.delete_smalt_button.config(state=tk.NORMAL)
            
    def combined_keyrelease_handler_smalt(self, event=None):
        self.update_representation_smalt()
        self.check_smalt_existence(event)
    
    #sauvegarde les modifications 
    #si un smalt existe avec ce nom, il le modifie, sinon il crée un nouveau smalt 
    def save_smalt(self):
        new_smalt = {label: self.smalt_entries[label].get() for label in self.smalt_labels if label not in ["type", "etat", "deplacement"]}
        new_smalt["type"] = self.smalt_type_var.get()
        new_smalt["etat"] = self.smalt_etat_var.get()
        new_smalt["deplacement"] = self.smalt_deplacement_var.get()
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "smalt"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
    
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        if self.smalt_row_index:
            old_nom = df.loc[self.smalt_row_index[0], 'nom']
            #Modifie le smalt existant
            df.loc[self.smalt_row_index[0]] = new_smalt
            print(f"Modification smalt dans {self.palier}_modifs_LHT.xlsx, feuille 'smalt'")
            # Met à jour les clefs dans la feuille clef
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                clef_df.loc[clef_df['appartenance'] == old_nom, 'appartenance'] = new_smalt['nom']

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()
                
                print(f"Les clefs dans {self.palier}_modifs_LHT.xlsx, feuille 'clef' ont été mises à jour avec le nouveau nom du smalt")
            else:
                messagebox.showwarning("Avertissement", "Feuille 'clef' non trouvée")
            self.smalt_row_index = []
        else:
            df = df.append(new_smalt, ignore_index=True)
            print(f"Nouveaux smaltx sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'smalt'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        
        print(f"Nouveeaux smalts sauvegardés dans {self.palier}_modifs_LHT.xlsx, feuille 'smalt'")

    #met à jour la représentation de la serrurerie du smalt, en allant chercher dans la feuille excel clefs les clefs qui appartiennent à ce smalt   
    def update_representation_smalt(self, event=None):
        representation = [ [0, 0]]

        smalt_nom = self.smalt_entries.get("nom", "").get()
        
        positions_rentre = []
        positions_sorti = []
        
        position_rentre_index = 0
        position_sorti_index = 0
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "clef"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        #compte le nombre de clefs 'rentre' et 'sorti'
        rentre_count = len(df[(df["appartenance"] == smalt_nom) & (df["penne"] == "rentre")])
        sorti_count = len(df[(df["appartenance"] == smalt_nom) & (df["penne"] == "sorti")])
        print(rentre_count)
        if sorti_count >= rentre_count:
            for i in range(sorti_count):
                representation.append([2, 0])
                representation.append([1, 0])
                positions_sorti.append((2*i+1, 1))
                positions_rentre.append((2*(i+1), 1))
            representation.append([2, 0])
        else:
            for i in range(rentre_count):
                representation.append([1, 0])
                representation.append([2, 0])
                positions_rentre.append((2*i+1, 1))
                positions_sorti.append((2*(i+1), 1))
        print(representation)
        print(positions_rentre)
        print(positions_sorti)
        #on parcourt toutes les clés dans la feuille Excel
        for _, row in df.iterrows():
            appartenance = row["appartenance"]
            if appartenance == smalt_nom:
                clef_nom = row["nom"]
                penne = row["penne"]
                cellule = row["cellule"]
                if clef_nom:
                    if penne == "rentre" and position_rentre_index < len(positions_rentre):
                        r, c = positions_rentre[position_rentre_index]
                        representation[r][c] = clef_nom
                        position_rentre_index += 1
                    elif penne == "sorti" and position_sorti_index < len(positions_sorti):
                        r, c = positions_sorti[position_sorti_index]
                        representation[r][c] = clef_nom
                        position_sorti_index += 1
                    else:
                        messagebox.showwarning()
                    
                    self.smalt_entries["cellule"].delete(0, tk.END)
                    self.smalt_entries["cellule"].insert(0, cellule)
                    
        #on met à jour le champ de saisie de la représentation
        self.smalt_entries["representation"].delete(0, tk.END)
        self.smalt_entries["representation"].insert(0, str(representation))

    #supprime un smalt existant    
    def delete_smalt(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "smalt"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.smalt_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            nom = df.loc[row_index[0], 'nom']
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            clef_sheet_name = "clef"
            if clef_sheet_name in book.sheetnames:
                clef_sheet = book[clef_sheet_name]
                clef_df = pd.DataFrame(clef_sheet.values)
                clef_df.columns = clef_df.iloc[0]
                clef_df = clef_df[1:]
                
                #Supprime les clefs qui ont comme appartenance le nom du smalt
                clef_df = clef_df[clef_df['appartenance'] != nom]

                if clef_sheet_name in book.sheetnames:
                    book.remove(clef_sheet)

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    writer.book = book
                    clef_df.to_excel(writer, sheet_name=clef_sheet_name, index=False)
                    writer.save()

            print(f"Smalt supprimé dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            #Réinitialise les champs de saisie
            for label in self.smalt_labels:
                self.smalt_entries[label].delete(0, tk.END)

    #crée l'onglet permettant de modifier des cellules
    def create_cellule_tab(self):
        self.cellule_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.cellule_tab, text="Cellule")
                
        self.cellule_labels = ["nom", "voisine", "position_x", "statut", "position_y", "representation"]
        self.cellule_entries = {}

        # Charger les données existantes pour "nom" et "voisine"
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "cellule"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            
            noms_existants = df['nom'].unique().tolist()  # Valeurs possibles pour "nom"
            voisines_existantes = df['voisine'].unique().tolist()  # Valeurs possibles pour "voisine"
        else:
            noms_existants = []
            voisines_existantes = []

        for i, label in enumerate(self.cellule_labels):
            ttk.Label(self.cellule_tab, text=label).grid(row=i, column=0, padx=10, pady=5)
            
            if label == "statut":
                self.statut_var = tk.StringVar(value="depart")
                type_menu = ttk.OptionMenu(self.cellule_tab, self.statut_var, "depart", "depart", "arrivee")
                type_menu.grid(row=i, column=1, padx=10, pady=5)
                ttk.Label(self.cellule_tab, text="(soit la cellule est dessinée au dessus du jeu de barre (arrivee), soit au dessous (depart))").grid(row=i, column=2, padx=10, pady=5)
            
            elif label == "nom":
                self.nom_combobox = ttk.Combobox(self.cellule_tab, values=noms_existants)
                self.nom_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.cellule_entries["nom"] = self.nom_combobox
                ttk.Label(self.cellule_tab, text="(Appuyer sur 'Entrée' après avoir sélectionné une cellule)").grid(row=i, column=2, padx=10, pady=5)
            
            elif label == "voisine":
                self.voisine_combobox = ttk.Combobox(self.cellule_tab, values=voisines_existantes)
                self.voisine_combobox.grid(row=i, column=1, padx=10, pady=5)
                self.cellule_entries["voisine"] = self.voisine_combobox
                self.add_voisine_button = ttk.Button(self.cellule_tab, text="Ajouter", command=self.add_voisine)
                self.add_voisine_button.grid(row=i, column=3, padx=10, pady=5)
                self.voisine_label = ttk.Label(self.cellule_tab, text="")
                self.voisine_label.grid(row=i, column=4, padx=10, pady=5)
                ttk.Label(self.cellule_tab, text="(écrire le nom de chaque cellule voisine et cliquer sur ajouter)").grid(row=i, column=2, padx=10, pady=5)
            
            else:
                entry = ttk.Entry(self.cellule_tab)
                entry.grid(row=i, column=1, padx=10, pady=5)
                self.cellule_entries[label] = entry

                if label == "representation":
                    ttk.Label(
                        self.cellule_tab,
                        text="(se remplit automatiquement, ne pas modifier)"
                    ).grid(row=i, column=2, padx=10, pady=5, sticky="w")

        self.rows_label = ttk.Label(self.cellule_tab, text="Nombre de lignes matrice (si création d'une nouvelle cellule):")
        self.rows_label.grid(row=len(self.cellule_labels), column=1, padx=10, pady=5)
        self.rows_entry = ttk.Entry(self.cellule_tab)
        self.rows_entry.grid(row=len(self.cellule_labels), column=2, padx=10, pady=5)

        self.create_matrix_button = ttk.Button(self.cellule_tab, text="Créer Cellule", command=self.create_cellule, state=tk.NORMAL)
        self.create_matrix_button.grid(row=len(self.cellule_labels), column=3, padx=10, pady=5)

        self.save_cellule_button = ttk.Button(self.cellule_tab, text="Sauvegarder", command=self.save_cellules)
        self.save_cellule_button.grid(row=len(self.cellule_labels), column=4, padx=10, pady=5)
        self.delete_cellule_button = ttk.Button(self.cellule_tab, text="Supprimer", command=self.delete_cellule, state=tk.DISABLED)
        self.delete_cellule_button.grid(row=len(self.cellule_labels) + 3, columnspan=3, padx=10, pady=10)

        self.add_row_button = ttk.Button(self.cellule_tab, text="Ajouter une ligne", command=self.add_row, state=tk.DISABLED)
        self.add_row_button.grid(row=len(self.cellule_labels), column=5, padx=10, pady=5)
        self.remove_row_button = ttk.Button(self.cellule_tab, text="Supprimer une ligne", command=self.remove_row, state=tk.DISABLED)
        self.remove_row_button.grid(row=len(self.cellule_labels), column=6, padx=10, pady=5)

        # Frame contenant le canvas et les scrollbars
        canvas_frame = ttk.Frame(self.cellule_tab)
        canvas_frame.grid(row=len(self.cellule_labels)+1, column=0, columnspan=6, padx=10, pady=10)

        # Canvas
        canvas_frame = ttk.Frame(self.cellule_tab)
        canvas_frame.grid(row=len(self.cellule_labels)+1, columnspan=6, padx=10, pady=10)

        self.canvas = tk.Canvas(canvas_frame, width=400, height=300, bg="white")
        self.canvas.grid(row=0, column=0)

        # Scrollbars
        h_scroll = ttk.Scrollbar(canvas_frame, orient="horizontal", command=self.canvas.xview)
        h_scroll.grid(row=1, column=0, sticky="ew")
        v_scroll = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")

        self.canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        # Zone de dessin logique
        self.canvas.config(scrollregion=(-50, -50, 2000, 1000))

        self.dessin = Dessin(self.canvas)

        # Créez un Canvas pour contenir les widgets
        self.canvas_frame = ttk.Frame(self.cellule_tab)
        self.canvas_frame.grid(row=len(self.cellule_labels) + 1, column=3, columnspan=5, sticky='n')
        self.canvas_matrix = tk.Canvas(self.canvas_frame, width=400, height=500)
        self.canvas_matrix.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Ajoutez des scrollbars au Canvas
        self.scrollbar_y = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas_matrix.yview)
        self.scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas_matrix.configure(yscrollcommand=self.scrollbar_y.set)

        self.scrollbar_x = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas_matrix.xview)
        self.scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas_matrix.configure(xscrollcommand=self.scrollbar_x.set)

        # Créez un Frame à l'intérieur du Canvas
        self.inner_frame = ttk.Frame(self.canvas_matrix)
        self.canvas_matrix.create_window((0, 0), window=self.inner_frame, anchor='nw')
        self.cellule_entries["nom"].bind("<Return>", self.load_cellule)
        self.load_and_draw_all_cells()
        
    #avant toute action sur l'onglet, dessine l'ensemble des cellule du tableau    
    def load_and_draw_all_cells(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "cellule"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        self.canvas.delete("all")
        
        for _, row in df.iterrows():
            nom = row["nom"]
            position_x = row["position_x"]
            position_y = row["position_y"]
            statut = row["statut"]
            representation_str = row["representation"]
            if representation_str is not None:
                representation = ast.literal_eval(representation_str)
            
            else:
                continue
            
            for i, line in enumerate(representation):
                for j, element in enumerate(line):
                    if element != 0:
                        x = position_x*5*self.cell_size + j * self.cell_size
                        y = i * self.cell_size + position_y*self.cell_size
                        if statut == 'depart':
                            self.draw_shape(element, x, y, self.cell_size)
                        elif statut ==  'arrivee':
                            self.draw_shape(element, x, y-(len(representation)-1)*self.cell_size, self.cell_size)
            self.canvas.create_text(position_x*5*self.cell_size + 30, position_y*self.cell_size - 5, text=nom, font=("Arial", 8))
            self.canvas.create_text(position_x*5*self.cell_size + 35, position_y*self.cell_size + 5, text=f"position_x : {position_x}, ")
            self.canvas.create_text(position_x*5*self.cell_size + 35, position_y*self.cell_size + 20, text=f"position_y : {position_y}")

    #permet d'ajouter les cellules voisines
    def add_voisine(self):
        voisine_name = self.voisine_combobox.get()
        if voisine_name:
            self.voisine_list.append(voisine_name)
            self.voisine_label.config(text=", ".join(self.voisine_list))
            self.voisine_combobox.set('')  # Réinitialise le combobox après ajout
        print(self.voisine_list)

    #dessine le composant de la matrice que l'on ajoute    
    def draw_shape(self, shape, x, y, size):
        if shape == 'disjoncteur':
            self.dessin.dessiner_disjoncteur(x + size/2, y, size)
        elif shape == 'contacteur':
            self.dessin.dessiner_contacteur(x + size/2, y, size)
        elif shape == 'tiroir':
            self.dessin.dessiner_tiroir(x + size/2, y, size)
        elif shape == 'fil_h':    
            self.dessin.dessiner_fil_h(x, y + size/2, size)
        elif shape == 'fil_v':
            self.dessin.dessiner_fil_v(x + size/2, y , size)
        elif shape == 'smalt':  
            self.dessin.dessiner_terre(x, y + size/2, size)        
        elif shape == 'fil_h_dep':
            self.dessin.dessiner_fil_h_dep(x, y , size) 
        elif shape == 'fil_h_arrivee':
            self.dessin.dessiner_fil_h_arrivee(x, y , size)       
        elif shape == 'fil_dep_smalt':
            self.dessin.dessiner_fil_dep_smalt(x, y , size)            
        elif shape == 'angle_haut_gauche':
           self.dessin.dessiner_angle_haut_gauche(x, y , size)           
        elif shape == 'angle_haut_droit':
           self.dessin.dessiner_angle_haut_droit(x, y , size)          
        elif shape == 'angle_bas_droit':
           self.dessin.dessiner_angle_bas_droit(x, y , size)           
        elif shape == 'angle_bas_gauche':
           self.dessin.dessiner_angle_bas_gauche(x, y , size)          
        elif shape == 'croisement':
           self.dessin.dessiner_croisement(x, y , size)            
        elif shape == 'source':
            self.dessin.dessiner_cercle(x + size/2 , y + size/2, size / 2)            
        elif shape == 'fusible':
            self.dessin.dessiner_fusible(x, y, size)
        elif shape == 'transformateur':
            rayon = size/4
            self.canvas.create_line(x + size/2, y, x + size/2, y+rayon, fill='black', width=2)
            self.dessin.dessiner_cercle( x + size/2, y+2*rayon, rayon)
            self.dessin.dessiner_cercle(x + size/2, y+3*rayon, rayon) 
        elif shape == 'coffret':
            self.canvas.create_rectangle(x, y, x + size, y + size, width = 2)
        elif shape == 'eclisse':
            self.dessin.dessiner_eclisse_presente(x + size/2, y, size)
        elif shape == 'boite_eclisse':
            self.canvas.create_line(x -size/2 +5 , y - size, x -size/2 +5 , y + size, fill='black', width=1)
            self.canvas.create_line(x -size/2 +8 , y - size, x -size/2 +8 , y + size, fill='black', width=1)
            self.canvas.create_line(x -size/2 , y - size , x , y - size, fill='black', width=2)
            self.canvas.create_line(x - size/2, y + size , x , y + size, fill='black', width=2)
            self.canvas.create_line(x - size/2 , y - size  , x -size/2, y + size, fill='black', width=2)

    #crée la nouvelle matrice    
    def create_matrix(self):
        rows = int(self.rows_entry.get())
        self.matrix = [[0 for _ in range(5)] for _ in range(rows)]
        self.update_matrix_ui(rows, self.cell_size, self.dessin)
        self.update_representation_cellule()

    #fonction qui ajoute une ligne à la matrice    
    def add_row(self):
        new_row = [0] * len(self.matrix[0])
        self.matrix.append(new_row)
        self.update_representation_cellule()  
        self.update_matrix_ui(len(self.matrix), self.cell_size, self.dessin)
        self.redraw_matrix()

    #fonction qui supprime la dernière ligne de la matrice 
    def remove_row(self):
        if len(self.matrix) > 0:
            print(self.matrix)
            self.matrix.pop()
            print()
            print()
            print(self.matrix)
            self.update_representation_cellule()  
            self.update_matrix_ui(len(self.matrix), self.cell_size, self.dessin)
            self.redraw_matrix()

    #vérifie si une cellule existe avec ce nom, et la charge le cas échéant    
    def load_cellule(self, event=None):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "cellule"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_load = self.cellule_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_load].tolist()
        
        if row_index:
            # Charge la cellule existante
            row = df.iloc[row_index[0] - 1]
                        
            self.cellule_entries["voisine"].delete(0, tk.END)
            self.cellule_entries["voisine"].insert(0, row["voisine"])

            self.cellule_entries["position_x"].delete(0, tk.END)
            self.cellule_entries["position_x"].insert(0, row["position_x"])

            self.cellule_entries["position_y"].delete(0, tk.END)
            self.cellule_entries["position_y"].insert(0, row["position_y"])

            self.cellule_entries["representation"].delete(0, tk.END)
            self.cellule_entries["representation"].insert(0, row["representation"])

            self.statut_var.set(row["statut"])
            
            representation = eval(row["representation"])
            self.matrix = representation

            self.position_x = int(row["position_x"])
            self.position_y = int(row["position_y"])
            self.update_matrix_ui(len(self.matrix), self.cell_size, self.dessin)
            self.redraw_matrix()
            
            self.delete_cellule_button.config(state=tk.NORMAL)
            self.add_row_button.config(state=tk.NORMAL)
            self.remove_row_button.config(state=tk.NORMAL)
            
            # Cache le bouton "Nombre de lignes" et l'entrée correspondante
            self.rows_label.grid_remove()
            self.rows_entry.grid_remove()
            self.create_matrix_button.config(state=tk.DISABLED)
        else:
            # Réinitialise les champs si le nom ne correspond pas à une cellule existante
            self.cellule_entries["voisine"].delete(0, tk.END)
            self.cellule_entries["position_x"].delete(0, tk.END)
            self.cellule_entries["position_y"].delete(0, tk.END)
            self.cellule_entries["representation"].delete(0, tk.END)
            
            # Réinitialise la matrice et la représentation
            self.matrix = []
            self.position_x = None
            self.update_matrix_ui(0, self.cell_size, self.dessin)
            self.canvas.delete("all")

            self.delete_cellule_button.config(state=tk.DISABLED)
            self.add_row_button.config(state=tk.DISABLED)
            self.remove_row_button.config(state=tk.DISABLED)
            
            # Affiche le bouton "Nombre de lignes" et l'entrée correspondante
            self.rows_label.grid()
            self.rows_entry.grid()
            self.create_matrix_button.config(state=tk.NORMAL)
        
    #crée une nouvelle cellule        
    def create_cellule(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "cellule"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        self.create_matrix()
        self.delete_cellule_button.config(state=tk.DISABLED)

    #fonction qui crée la matrice avec pour case, un bandeau avec les possibilités  
    def update_matrix_ui(self, rows, cell_size, dessin, offset_x=0):
        options = [0, 0, "disjoncteur", "contacteur", "tiroir", "fusible", "smalt", "source", "fil_h", "fil_v", "transformateur",
                   "coffret", "boite_eclisse", "eclisse", "angle_haut_gauche", "angle_haut_droit", "angle_bas_gauche", "angle_bas_droit", "fil_dep_smalt",
                   "fil_h_dep", "fil_h_arrivee", "croisement"]

        # Détruisez les widgets existants dans inner_frame
        for widget in self.inner_frame.winfo_children():
            widget.destroy()

        for i in range(rows):
            for j in range(5):
                var = tk.StringVar(value=self.matrix[i][j])
                dropdown = ttk.OptionMenu(self.inner_frame, var, *options, command=lambda v, r=i, c=j: self.update_matrix_and_draw(r, c, v, dessin, cell_size, offset_x))
                dropdown.grid(row=i, column=j, padx=5, pady=5)

        # Mettez à jour la taille du Canvas
        self.inner_frame.update_idletasks()
        self.canvas_matrix.config(scrollregion=self.canvas_matrix.bbox("all"))

    #redessine la matrice
    def redraw_matrix(self):
        self.canvas.delete("all")
        
        for i in range(len(self.matrix)):
            for j in range(5):
                if self.matrix[i][j] != 0:
                    x = j * self.cell_size 
                    y = i * self.cell_size
                    self.draw_shape(self.matrix[i][j], x, y, self.cell_size)

    #permet de recentrer le dessin sur les nouveaux éléments dessinés
    def center_canvas_on_new_elements(self, row, col, cell_size):
        x = col * cell_size
        y = row * cell_size
        self.canvas.xview_moveto(x / self.canvas.winfo_width())
        self.canvas.yview_moveto(y / self.canvas.winfo_height())

    #met à jour la matrice et le dessin            
    def update_matrix_and_draw(self, row, col, value, dessin, cell_size, offset_x=0):
        self.matrix[row][col] = value
        print(self.matrix)
        self.update_representation_cellule()
        x = col * self.cell_size + offset_x
        y = row * self.cell_size
        self.canvas.delete("all")
        for i in range(len(self.matrix)):
            for j in range(5):
                if self.matrix[i][j] != 0:
                    x = j * self.cell_size + offset_x
                    y = i * self.cell_size
                    self.draw_shape(self.matrix[i][j], x, y, self.cell_size)
        self.center_canvas_on_new_elements(row, col, cell_size)

    #met à jour la matrice de la représentation de la cellule   
    def update_representation_cellule(self):
        representation = [[0, 0, 0, 0, 0] for _ in range(len(self.matrix))]
        
        for i in range(len(self.matrix)):
            for j in range(5):
                if self.matrix[i][j] != 0:
                    representation[i][j] = self.matrix[i][j]
        
        self.cellule_entries["representation"].delete(0, tk.END)
        self.cellule_entries["representation"].insert(0, str(representation))

    #supprime une cellule existante    
    def delete_cellule(self):
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "cellule"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")
            return
        
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        
        nom_to_delete = self.cellule_entries["nom"].get()
        row_index = df.index[df['nom'] == nom_to_delete].tolist()
        
        if row_index:
            #Supprime la ligne existante
            df = df.drop(row_index[0])
            
            book.remove(sheet)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()
            
            print(f"Cellule supprimée dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
            
            #Réinitialise les champs de saisie
            for label in self.cellule_labels:
                self.cellule_entries[label].delete(0, tk.END)
            
            self.delete_cellule_button.config(state=tk.DISABLED)
            
            self.canvas.delete("all")
        else:
            messagebox.showwarning("Avertissement", "Cellule non trouvée")
        self.load_and_draw_all_cells()

    #sauvegarde les modifications 
    #si une cellule existe avec ce nom, il la modifie, sinon il crée une nouvelle cellule     
    def save_cellules(self):
        new_cellule = {}
        
        new_cellule = {}
        for label in self.cellule_labels:
            print(label)
            if label == "voisine":
                value = self.voisine_list  
                new_cellule[label] = value
            elif label == "position_x" or label == "position_y":
                value = self.cellule_entries[label].get()
                new_cellule[label] = int(value)
            elif label == "statut":
                new_cellule[label] = self.statut_var.get()
            else:
                value = self.cellule_entries[label].get()
                new_cellule[label] = value 
        
        excel_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        sheet_name = "cellule"
        
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            messagebox.showwarning("Avertissement", "Fichier Excel non trouvé")

        
        if sheet_name in book.sheetnames:
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
        
        #Recherche si la cellule existe déjà
        nom_to_save = new_cellule["nom"]
        row_index = df.index[df['nom'] == nom_to_save].tolist()
        
        if row_index:
            #Remplace la ligne existante
            for label in self.cellule_labels:
                df.at[row_index[0], label] = new_cellule[label]
            print(f"modification cellule {nom_to_save} dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
        else:
            #Ajoute une nouvelle ligne
            df = df.append(new_cellule, ignore_index=True)
            print(f"Nouvelles cellules sauvegardées dans {self.palier}_modifs_LHT.xlsx, feuille 'cellule'")
        
        if sheet_name in book.sheetnames:
            book.remove(sheet)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        self.load_and_draw_all_cells()

    #fonction exécutée quand on choisit le palier pour afficher le tableau du palier choisi
    #Permet de visualiser le tableau : pratique pour se repérer pour effectuer les modifications
    def tableauLHT_tab(self):
        etapes = EtapesChoisies()
        initialisation = InitialisationCellules(f"{self.palier}_modifs")
        tableau = ExecutionLHT(self.window)
        tableau.tableauLHT(f"{self.palier}_modifs", etapes)

    #fonction pour transférer les modifications faites sur un fichier Excel à part sur le fichier Excel principal
    def create_update_palier_tab(self):
        self.update_palier_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.update_palier_tab, text="Mettre à jour Excel principal")
        
        ttk.Label(self.update_palier_tab, text="Mettre à jour le fichier Excel principal pour ajouter les dernières modifications").pack(padx=10, pady=10)
        ttk.Label(self.update_palier_tab, text=f"Attention ! Assurez d'être sur le palier que vous voulez mettre à jour !").pack(padx=10, pady=10)
        
        self.update_button = ttk.Button(self.update_palier_tab, text="Mettre à jour", command=self.update_palier)
        self.update_button.pack(padx=10, pady=10)

    #liée au boutons de mise à jour du fichier Excel principal
    def update_palier(self):
        modifs_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        palier_path = f"dataLHT/{self.palier}_LHT.xlsx"
        
        modifs_book = load_workbook(modifs_path)
        palier_book = load_workbook(palier_path)
        
        #Copie les feuilles de palier_modifs vers palier
        for sheet_name in modifs_book.sheetnames:
            if sheet_name in palier_book.sheetnames:
                palier_book.remove(palier_book[sheet_name])
            modifs_sheet = modifs_book[sheet_name]
            palier_book.create_sheet(sheet_name)
            new_sheet = palier_book[sheet_name]
            
            for row in modifs_sheet.iter_rows(values_only=True):
                new_sheet.append(row)
        
        palier_book.save(palier_path)
        
        messagebox.showinfo("Mise à jour", f"Le fichier {palier_path} a été mis à jour avec succès à partir de {modifs_path}")

    #fonction pour revenir à l'état avant les modifs
    def create_retour_palier_tab(self):
        self.retour_palier_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.retour_palier_tab, text="Revenir état avant modifs")
        
        ttk.Label(self.retour_palier_tab, text="Si les modifications n'aboutissent pas, revenir à l'état précédent pour tester d'autres modifs").pack(padx=10, pady=10)
        ttk.Label(self.retour_palier_tab, text=f"Attention ! Assurez d'être sur le palier que vous voulez mettre à jour !").pack(padx=10, pady=10)
        
        self.retour_button = ttk.Button(self.retour_palier_tab, text="Mettre à jour", command=self.retour_palier)
        self.retour_button.pack(padx=10, pady=10)

    #liée au bouton de mise à jour du fichier Excel des modifications
    def retour_palier(self):
        modifs_path = f"dataLHT/{self.palier}_modifs_LHT.xlsx"
        palier_path = f"dataLHT/{self.palier}_LHT.xlsx"
        
        palier_book = load_workbook(palier_path)
        modifs_book = load_workbook(modifs_path)
        
        #Copie les feuilles de palier_modifs vers palier
        for sheet_name in palier_book.sheetnames:
            if sheet_name in modifs_book.sheetnames:
                modifs_book.remove(modifs_book[sheet_name])
            palier_sheet = palier_book[sheet_name]
            modifs_book.create_sheet(sheet_name)
            new_sheet = modifs_book[sheet_name]
            
            for row in palier_sheet.iter_rows(values_only=True):
                new_sheet.append(row)
        
        modifs_book.save(modifs_path)
        
        messagebox.showinfo("Mise à jour", "Le fichier retrouve bien son état avant les modifications")

#lance l'interface
def main():
    root = tk.Toplevel(root)
    app = ModifsExcel(root)
    root.mainloop()
if __name__ == "__main__":
    main()
    